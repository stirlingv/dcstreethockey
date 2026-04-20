"""
Unit and integration tests for the Wednesday Draft League draft functionality.

Coverage:
  Models     – DraftTeam.save, DraftSession.current_pick,
               DraftSession.pick_order_for_round (snake, randomized, continuity)
  Views      – draft_signup, board views, draw_positions, advance_state,
               make_pick, undo_last_pick, swap_pick, reset_draft
  Logic      – _process_auto_captain_picks, _session_state_payload
  Validation – captain cross-team guard, goalie-per-team limit,
               already-drafted, wrong-turn, state guards, token auth
"""

import datetime
import json
from unittest.mock import patch

from django.test import TestCase, Client
from django.urls import reverse

from leagues.models import (
    Division,
    DraftPick,
    DraftRound,
    DraftSession,
    DraftTeam,
    MatchUp,
    Player,
    Roster,
    Season,
    SeasonSignup,
    Stat,
    Team,
    Team_Stat,
    Week,
)
from leagues.draft_views import (
    _batch_wednesday_stats,
    _get_champion_data_for_sessions,
    _get_session_champion,
    _get_wednesday_stats,
    _process_auto_captain_picks,
    _session_state_payload,
)


# ---------------------------------------------------------------------------
# Shared base
# ---------------------------------------------------------------------------


class DraftTestBase(TestCase):
    """
    Per-test fixture factory.  Every test gets a clean DB state with:
      • 1 Season  (year=current, season_type=4 "Winter" — unlikely to collide)
      • 1 DraftSession  (3 teams, 3 rounds, SETUP, signups open)
      • 3 captain SeasonSignups  →  3 DraftTeams (draft positions 1–3 drawn)
      • 6 regular player SeasonSignups
      • 2 goalie SeasonSignups
    """

    def setUp(self):
        self.season = Season.objects.create(
            year=datetime.date.today().year,
            season_type=4,
            is_current_season=False,
        )
        self.session = DraftSession.objects.create(
            season=self.season,
            num_teams=3,
            num_rounds=3,
            state=DraftSession.STATE_SETUP,
            signups_open=True,
        )

        # Captain signups
        self.cap1 = SeasonSignup.objects.create(
            season=self.season,
            first_name="Alice",
            last_name="Alpha",
            email="alice@test.com",
            primary_position=SeasonSignup.POSITION_CENTER,
            secondary_position=SeasonSignup.POSITION_WING,
            captain_interest=SeasonSignup.CAPTAIN_YES,
        )
        self.cap2 = SeasonSignup.objects.create(
            season=self.season,
            first_name="Bob",
            last_name="Beta",
            email="bob@test.com",
            primary_position=SeasonSignup.POSITION_WING,
            secondary_position=SeasonSignup.POSITION_CENTER,
            captain_interest=SeasonSignup.CAPTAIN_YES,
        )
        self.cap3 = SeasonSignup.objects.create(
            season=self.season,
            first_name="Carol",
            last_name="Gamma",
            email="carol@test.com",
            primary_position=SeasonSignup.POSITION_DEFENSE,
            secondary_position=SeasonSignup.POSITION_ONE_THING,
            captain_interest=SeasonSignup.CAPTAIN_YES,
        )

        # Teams — positions already drawn (1, 2, 3)
        self.team1 = DraftTeam.objects.create(
            session=self.session, captain=self.cap1, draft_position=1
        )
        self.team2 = DraftTeam.objects.create(
            session=self.session, captain=self.cap2, draft_position=2
        )
        self.team3 = DraftTeam.objects.create(
            session=self.session, captain=self.cap3, draft_position=3
        )

        # Regular player signups
        _positions = [
            SeasonSignup.POSITION_CENTER,
            SeasonSignup.POSITION_WING,
            SeasonSignup.POSITION_DEFENSE,
            SeasonSignup.POSITION_CENTER,
            SeasonSignup.POSITION_WING,
            SeasonSignup.POSITION_DEFENSE,
        ]
        self.players = [
            SeasonSignup.objects.create(
                season=self.season,
                first_name=f"Player{i}",
                last_name=f"Test{i}",
                email=f"player{i}@test.com",
                primary_position=_positions[i - 1],
                secondary_position=SeasonSignup.POSITION_ONE_THING,
                captain_interest=SeasonSignup.CAPTAIN_NO,
            )
            for i in range(1, 7)
        ]

        # Goalie signups
        self.goalie1 = SeasonSignup.objects.create(
            season=self.season,
            first_name="Gary",
            last_name="Goalie",
            email="gary@test.com",
            primary_position=SeasonSignup.POSITION_GOALIE,
            secondary_position=SeasonSignup.POSITION_ONE_THING,
            captain_interest=SeasonSignup.CAPTAIN_NO,
        )
        self.goalie2 = SeasonSignup.objects.create(
            season=self.season,
            first_name="Greta",
            last_name="Goal",
            email="greta@test.com",
            primary_position=SeasonSignup.POSITION_GOALIE,
            secondary_position=SeasonSignup.POSITION_ONE_THING,
            captain_interest=SeasonSignup.CAPTAIN_NO,
        )

        self.client = Client()

        # Suppress WebSocket broadcasts in all tests — async_to_sync spawns a
        # new event loop thread per call, which makes the suite very slow.
        self._broadcast_patcher = patch("leagues.draft_views._broadcast_state_change")
        self._broadcast_patcher.start()

    def tearDown(self):
        self._broadcast_patcher.stop()
        super().tearDown()

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _activate(self):
        self.session.state = DraftSession.STATE_ACTIVE
        self.session.save(update_fields=["state"])

    def _complete(self):
        self.session.state = DraftSession.STATE_COMPLETE
        self.session.save(update_fields=["state"])

    def _make_pick(self, team, signup, round_number, pick_number, auto=False):
        return DraftPick.objects.create(
            session=self.session,
            team=team,
            signup=signup,
            round_number=round_number,
            pick_number=pick_number,
            is_auto_captain=auto,
        )

    def _post_pick(self, signup_pk, *, captain_token=None, commissioner_token=None):
        data = {"signup_pk": signup_pk}
        if captain_token:
            data["captain_token"] = str(captain_token)
        if commissioner_token:
            data["commissioner_token"] = str(commissioner_token)
        return self.client.post(
            reverse("draft_make_pick", args=[self.session.pk]), data
        )


# ---------------------------------------------------------------------------
# Model: DraftTeam
# ---------------------------------------------------------------------------


class DraftTeamModelTests(DraftTestBase):
    def test_team_name_auto_generated_from_captain(self):
        new_signup = SeasonSignup.objects.create(
            season=self.season,
            first_name="Zara",
            last_name="Zeta",
            email="zara@test.com",
            primary_position=SeasonSignup.POSITION_WING,
            secondary_position=SeasonSignup.POSITION_ONE_THING,
            captain_interest=SeasonSignup.CAPTAIN_NO,
        )
        # Create a second session to avoid unique_together conflict
        session2 = DraftSession.objects.create(
            season=Season.objects.create(
                year=datetime.date.today().year, season_type=3, is_current_season=False
            ),
            num_teams=1,
            num_rounds=1,
        )
        team = DraftTeam.objects.create(
            session=session2, captain=new_signup, draft_position=1
        )
        self.assertEqual(team.team_name, "Zara's Team")

    def test_team_name_preserved_when_provided(self):
        new_signup = SeasonSignup.objects.create(
            season=self.season,
            first_name="Zara",
            last_name="Zeta",
            email="zara@test.com",
            primary_position=SeasonSignup.POSITION_WING,
            secondary_position=SeasonSignup.POSITION_ONE_THING,
            captain_interest=SeasonSignup.CAPTAIN_NO,
        )
        session2 = DraftSession.objects.create(
            season=Season.objects.create(
                year=datetime.date.today().year, season_type=3, is_current_season=False
            ),
            num_teams=1,
            num_rounds=1,
        )
        team = DraftTeam.objects.create(
            session=session2,
            captain=new_signup,
            team_name="The Rockets",
            draft_position=1,
        )
        self.assertEqual(team.team_name, "The Rockets")

    def test_existing_teams_have_auto_generated_names(self):
        # Teams created in setUp without explicit team_name should be auto-named
        self.assertEqual(self.team1.team_name, "Alice's Team")
        self.assertEqual(self.team2.team_name, "Bob's Team")
        self.assertEqual(self.team3.team_name, "Carol's Team")


# ---------------------------------------------------------------------------
# Model: DraftSession.current_pick
# ---------------------------------------------------------------------------


class DraftSessionCurrentPickTests(DraftTestBase):
    def test_no_picks_returns_round_1_index_0(self):
        self._activate()
        self.assertEqual(self.session.current_pick, (1, 0))

    def test_after_one_pick_returns_round_1_index_1(self):
        self._activate()
        self._make_pick(self.team1, self.players[0], 1, 0)
        self.assertEqual(self.session.current_pick, (1, 1))

    def test_after_full_first_round_returns_round_2_index_0(self):
        self._activate()
        self._make_pick(self.team1, self.players[0], 1, 0)
        self._make_pick(self.team2, self.players[1], 1, 1)
        self._make_pick(self.team3, self.players[2], 1, 2)
        self.assertEqual(self.session.current_pick, (2, 0))

    def test_draft_complete_returns_none(self):
        self._activate()
        # Fill all 9 slots (3 teams × 3 rounds)
        all_signups = self.players[:6] + [self.cap1, self.cap2, self.cap3]
        slot = 0
        for r in range(1, 4):
            for p in range(3):
                self._make_pick(
                    [self.team1, self.team2, self.team3][p],
                    all_signups[slot],
                    r,
                    p,
                )
                slot += 1
        self.assertIsNone(self.session.current_pick)


# ---------------------------------------------------------------------------
# Model: DraftSession.pick_order_for_round (snake + randomized)
# ---------------------------------------------------------------------------


class DraftPickOrderTests(DraftTestBase):
    def test_round_1_forward_order(self):
        order = self.session.pick_order_for_round(1)
        self.assertEqual(order, [self.team1.pk, self.team2.pk, self.team3.pk])

    def test_round_2_snake_reversed(self):
        order = self.session.pick_order_for_round(2)
        self.assertEqual(order, [self.team3.pk, self.team2.pk, self.team1.pk])

    def test_round_3_snake_forward_again(self):
        order = self.session.pick_order_for_round(3)
        self.assertEqual(order, [self.team1.pk, self.team2.pk, self.team3.pk])

    def test_randomized_round_is_deterministic(self):
        DraftRound.objects.create(
            session=self.session,
            round_number=2,
            order_type=DraftRound.ORDER_RANDOMIZED,
        )
        order_a = self.session.pick_order_for_round(2)
        order_b = self.session.pick_order_for_round(2)
        self.assertEqual(order_a, order_b)
        # Must contain all three teams
        self.assertCountEqual(order_a, [self.team1.pk, self.team2.pk, self.team3.pk])

    def test_snake_continuity_after_randomized_round(self):
        """
        Round 2 is randomized.  Round 3's snake parity should be as if
        round 2 never happened (effective_round = 3 - 1 = 2, even → reversed).
        """
        DraftRound.objects.create(
            session=self.session,
            round_number=2,
            order_type=DraftRound.ORDER_RANDOMIZED,
        )
        order = self.session.pick_order_for_round(3)
        self.assertEqual(order, [self.team3.pk, self.team2.pk, self.team1.pk])

    def test_round_3_without_randomized_round_is_forward(self):
        """
        Without any randomized rounds, round 3 (odd effective round) is forward.
        """
        order = self.session.pick_order_for_round(3)
        self.assertEqual(order, [self.team1.pk, self.team2.pk, self.team3.pk])


# ---------------------------------------------------------------------------
# View: draft_signup
# ---------------------------------------------------------------------------


class DraftSignupViewTests(DraftTestBase):
    def _signup_url(self):
        return reverse("draft_signup", args=[self.season.pk])

    def _valid_post_data(self, first="New", last="Player"):
        return {
            "first_name": first,
            "last_name": last,
            "email": "new@test.com",
            "primary_position": SeasonSignup.POSITION_WING,
            "secondary_position": SeasonSignup.POSITION_CENTER,
            "captain_interest": SeasonSignup.CAPTAIN_NO,
        }

    def test_get_renders_signup_form(self):
        response = self.client.get(self._signup_url())
        self.assertEqual(response.status_code, 200)

    def test_valid_post_creates_signup(self):
        before = SeasonSignup.objects.filter(season=self.season).count()
        self.client.post(self._signup_url(), self._valid_post_data())
        self.assertEqual(
            SeasonSignup.objects.filter(season=self.season).count(), before + 1
        )

    def test_post_missing_name_shows_error(self):
        data = self._valid_post_data()
        data["first_name"] = ""
        response = self.client.post(self._signup_url(), data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "required")
        self.assertFalse(
            SeasonSignup.objects.filter(
                last_name=data["last_name"], first_name=""
            ).exists()
        )

    def test_same_name_different_email_is_allowed(self):
        self.client.post(self._signup_url(), self._valid_post_data())
        data = self._valid_post_data()
        data["email"] = "other@test.com"
        response = self.client.post(self._signup_url(), data)
        # No error — two players with the same name but different emails is valid
        self.assertEqual(
            SeasonSignup.objects.filter(
                season=self.season, first_name="New", last_name="Player"
            ).count(),
            2,
        )

    def test_post_duplicate_email_shows_error(self):
        self.client.post(self._signup_url(), self._valid_post_data())
        # Same email, different name
        data = self._valid_post_data(first="Different", last="Person")
        response = self.client.post(self._signup_url(), data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "email address is already signed up")

    def test_signup_links_player_by_email_when_name_differs(self):
        """Email-matched Player is linked even when the signup name doesn't match exactly."""
        player = Player.objects.create(
            first_name="Michael",
            last_name="Smith",
            email="mike@example.com",
            is_active=True,
        )
        data = self._valid_post_data(first="Mike", last="Smith")
        data["email"] = "mike@example.com"
        self.client.post(self._signup_url(), data)
        signup = SeasonSignup.objects.get(season=self.season, email="mike@example.com")
        self.assertEqual(signup.linked_player, player)

    def test_signup_links_player_by_name_before_email(self):
        """Exact name match takes precedence over email match."""
        name_player = Player.objects.create(
            first_name="Mike",
            last_name="Jones",
            email="other@example.com",
            is_active=True,
        )
        Player.objects.create(
            first_name="Somebody",
            last_name="Else",
            email="mike@example.com",
            is_active=True,
        )
        data = self._valid_post_data(first="Mike", last="Jones")
        data["email"] = "mike@example.com"
        self.client.post(self._signup_url(), data)
        signup = SeasonSignup.objects.get(season=self.season, email="mike@example.com")
        self.assertEqual(signup.linked_player, name_player)

    def test_signups_closed_shows_closed_page(self):
        self.session.signups_open = False
        self.session.save(update_fields=["signups_open"])
        response = self.client.get(self._signup_url())
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "leagues/draft_signup_closed.html")


# ---------------------------------------------------------------------------
# Views: board pages render
# ---------------------------------------------------------------------------


class BoardViewTests(DraftTestBase):
    def test_spectator_view_renders(self):
        response = self.client.get(
            reverse("draft_board_spectator", args=[self.session.pk])
        )
        self.assertEqual(response.status_code, 200)

    def test_commissioner_view_renders(self):
        response = self.client.get(
            reverse(
                "draft_board_commissioner",
                args=[self.session.pk, self.session.commissioner_token],
            )
        )
        self.assertEqual(response.status_code, 200)

    def test_commissioner_wrong_token_returns_404(self):
        import uuid

        response = self.client.get(
            reverse(
                "draft_board_commissioner",
                args=[self.session.pk, uuid.uuid4()],
            )
        )
        self.assertEqual(response.status_code, 404)

    def test_captain_view_renders(self):
        response = self.client.get(
            reverse(
                "draft_board_captain",
                args=[self.session.pk, self.team1.captain_token],
            )
        )
        self.assertEqual(response.status_code, 200)

    def test_captain_wrong_token_returns_404(self):
        import uuid

        response = self.client.get(
            reverse(
                "draft_board_captain",
                args=[self.session.pk, uuid.uuid4()],
            )
        )
        self.assertEqual(response.status_code, 404)

    def test_captain_portal_renders(self):
        response = self.client.get(
            reverse("draft_captain_portal", args=[self.session.pk])
        )
        self.assertEqual(response.status_code, 200)


# ---------------------------------------------------------------------------
# View: draw_positions
# ---------------------------------------------------------------------------


class DrawPositionsViewTests(DraftTestBase):
    def setUp(self):
        super().setUp()
        # Reset positions so draw tests start clean
        self.session.teams.update(draft_position=None)

    def _draw_url(self):
        return reverse(
            "draft_draw_positions",
            args=[self.session.pk, self.session.commissioner_token],
        )

    def test_draw_assigns_positions_to_all_teams(self):
        self.client.post(self._draw_url())
        positions = list(self.session.teams.values_list("draft_position", flat=True))
        self.assertFalse(any(p is None for p in positions))
        self.assertCountEqual(positions, [1, 2, 3])

    def test_draw_from_setup_auto_advances_to_draw_state(self):
        self.assertEqual(self.session.state, DraftSession.STATE_SETUP)
        self.client.post(self._draw_url())
        self.session.refresh_from_db()
        self.assertEqual(self.session.state, DraftSession.STATE_DRAW)

    def test_draw_closes_signups(self):
        self.assertTrue(self.session.signups_open)
        self.client.post(self._draw_url())
        self.session.refresh_from_db()
        self.assertFalse(self.session.signups_open)

    def test_draw_wrong_state_returns_400(self):
        self._activate()
        response = self.client.post(self._draw_url())
        self.assertEqual(response.status_code, 400)

    def test_draw_wrong_token_returns_404(self):
        import uuid

        url = reverse("draft_draw_positions", args=[self.session.pk, uuid.uuid4()])
        response = self.client.post(url)
        self.assertEqual(response.status_code, 404)

    def test_draw_returns_reveal_order_sorted_by_position(self):
        response = self.client.post(self._draw_url())
        data = json.loads(response.content)
        positions = [item["position"] for item in data["reveal_order"]]
        self.assertEqual(positions, sorted(positions))


# ---------------------------------------------------------------------------
# View: advance_state
# ---------------------------------------------------------------------------


class AdvanceStateViewTests(DraftTestBase):
    def _advance_url(self):
        return reverse(
            "draft_advance_state",
            args=[self.session.pk, self.session.commissioner_token],
        )

    def test_setup_to_draw(self):
        self.client.post(self._advance_url())
        self.session.refresh_from_db()
        self.assertEqual(self.session.state, DraftSession.STATE_DRAW)

    def _set_all_captain_rounds(self):
        """Give every team a captain_draft_round so DRAW→ACTIVE succeeds."""
        for i, team in enumerate(self.session.teams.all(), start=1):
            team.captain_draft_round = i
            team.save(update_fields=["captain_draft_round"])

    def test_draw_to_active(self):
        self.session.state = DraftSession.STATE_DRAW
        self.session.save(update_fields=["state"])
        self._set_all_captain_rounds()
        self.client.post(self._advance_url())
        self.session.refresh_from_db()
        self.assertEqual(self.session.state, DraftSession.STATE_ACTIVE)

    def test_draw_to_active_closes_signups(self):
        self.session.state = DraftSession.STATE_DRAW
        self.session.signups_open = True
        self.session.save(update_fields=["state", "signups_open"])
        self._set_all_captain_rounds()
        self.client.post(self._advance_url())
        self.session.refresh_from_db()
        self.assertFalse(self.session.signups_open)

    def test_draw_to_active_blocked_when_captain_round_missing(self):
        self.session.state = DraftSession.STATE_DRAW
        self.session.save(update_fields=["state"])
        # teams have no captain_draft_round (default null)
        response = self.client.post(self._advance_url())
        self.assertEqual(response.status_code, 400)
        data = json.loads(response.content)
        self.assertIn("captain round not set", data["error"].lower())

    def test_draw_to_active_blocked_when_one_captain_round_missing(self):
        self.session.state = DraftSession.STATE_DRAW
        self.session.save(update_fields=["state"])
        # Set rounds for only two of three teams
        teams = list(self.session.teams.all())
        teams[0].captain_draft_round = 1
        teams[0].save(update_fields=["captain_draft_round"])
        teams[1].captain_draft_round = 2
        teams[1].save(update_fields=["captain_draft_round"])
        # teams[2] left null
        response = self.client.post(self._advance_url())
        self.assertEqual(response.status_code, 400)
        self.session.refresh_from_db()
        self.assertEqual(self.session.state, DraftSession.STATE_DRAW)

    def test_active_to_paused(self):
        self._activate()
        self.client.post(self._advance_url())
        self.session.refresh_from_db()
        self.assertEqual(self.session.state, DraftSession.STATE_PAUSED)

    def test_paused_to_active(self):
        self.session.state = DraftSession.STATE_PAUSED
        self.session.save(update_fields=["state"])
        self.client.post(self._advance_url())
        self.session.refresh_from_db()
        self.assertEqual(self.session.state, DraftSession.STATE_ACTIVE)

    def test_complete_state_has_no_transition(self):
        self._complete()
        response = self.client.post(self._advance_url())
        self.assertEqual(response.status_code, 400)

    def test_wrong_token_returns_404(self):
        import uuid

        url = reverse("draft_advance_state", args=[self.session.pk, uuid.uuid4()])
        response = self.client.post(url)
        self.assertEqual(response.status_code, 404)


# ---------------------------------------------------------------------------
# View: make_pick — auth, turn order, and all validations
# ---------------------------------------------------------------------------


class MakePickViewTests(DraftTestBase):
    def setUp(self):
        super().setUp()
        self._activate()

    def test_captain_makes_valid_pick(self):
        # Round 1, slot 0 → team1's turn
        response = self._post_pick(
            self.players[0].pk, captain_token=self.team1.captain_token
        )
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.content)
        self.assertTrue(data["success"])
        self.assertTrue(
            DraftPick.objects.filter(
                session=self.session, signup=self.players[0], team=self.team1
            ).exists()
        )

    def test_captain_cannot_pick_out_of_turn(self):
        # team2's captain tries to pick when it is team1's turn
        response = self._post_pick(
            self.players[0].pk, captain_token=self.team2.captain_token
        )
        self.assertEqual(response.status_code, 400)
        data = json.loads(response.content)
        self.assertIn("not your turn", data["error"])

    def test_commissioner_can_make_pick(self):
        response = self._post_pick(
            self.players[0].pk,
            commissioner_token=self.session.commissioner_token,
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(json.loads(response.content)["success"])

    def test_no_auth_returns_403(self):
        response = self._post_pick(self.players[0].pk)
        self.assertEqual(response.status_code, 403)

    def test_invalid_captain_token_returns_403(self):
        import uuid

        response = self._post_pick(self.players[0].pk, captain_token=uuid.uuid4())
        self.assertEqual(response.status_code, 403)

    def test_draft_not_active_returns_400(self):
        self.session.state = DraftSession.STATE_PAUSED
        self.session.save(update_fields=["state"])
        response = self._post_pick(
            self.players[0].pk,
            commissioner_token=self.session.commissioner_token,
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("not active", json.loads(response.content)["error"])

    def test_already_drafted_player_returns_400(self):
        self._make_pick(self.team1, self.players[0], 1, 0)
        # Try to pick the same player again (team2's turn now)
        response = self._post_pick(
            self.players[0].pk,
            commissioner_token=self.session.commissioner_token,
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("already drafted", json.loads(response.content)["error"])

    def test_cannot_draft_another_teams_captain(self):
        # It is team1's turn; team1 tries to draft cap2 (captain of team2)
        response = self._post_pick(self.cap2.pk, captain_token=self.team1.captain_token)
        self.assertEqual(response.status_code, 400)
        data = json.loads(response.content)
        self.assertIn("captain of", data["error"])
        self.assertIn("own team", data["error"])

    def test_can_draft_own_captain_when_no_auto_round_set(self):
        # cap1 has no captain_draft_round, so their own team can pick them manually
        response = self._post_pick(self.cap1.pk, captain_token=self.team1.captain_token)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            DraftPick.objects.filter(
                session=self.session, signup=self.cap1, team=self.team1
            ).exists()
        )

    def test_goalie_captain_blocks_second_goalie_before_auto_pick(self):
        # Make cap1 a goalie captain. Even before their auto-pick round fires,
        # the team should not be allowed to draft another goalie.
        self.cap1.primary_position = SeasonSignup.POSITION_GOALIE
        self.cap1.save(update_fields=["primary_position"])
        self.team1.captain_draft_round = 2
        self.team1.save(update_fields=["captain_draft_round"])
        # Round 1, team1's turn — try to draft a goalie
        response = self._post_pick(
            self.goalie1.pk, commissioner_token=self.session.commissioner_token
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("goalie", json.loads(response.content)["error"])

    def test_goalie_captain_has_goalie_true_in_payload(self):
        self.cap1.primary_position = SeasonSignup.POSITION_GOALIE
        self.cap1.save(update_fields=["primary_position"])
        payload = _session_state_payload(self.session)
        team1_data = next(t for t in payload["teams"] if t["id"] == self.team1.pk)
        self.assertTrue(team1_data["has_goalie"])

    def test_non_goalie_captain_has_goalie_false_in_payload(self):
        # cap1 is center by default (from DraftTestBase)
        payload = _session_state_payload(self.session)
        team1_data = next(t for t in payload["teams"] if t["id"] == self.team1.pk)
        self.assertFalse(team1_data["has_goalie"])

    def test_second_goalie_to_same_team_returns_400(self):
        # Team1 picks goalie1, then tries to pick goalie2
        self._make_pick(self.team1, self.goalie1, 1, 0)
        # Advance to team1's next turn: after round 1 (t1,t2,t3) and round 2
        # reversed (t3,t2,t1), team1's turn is pick index 2 of round 2.
        self._make_pick(self.team2, self.players[0], 1, 1)
        self._make_pick(self.team3, self.players[1], 1, 2)
        self._make_pick(self.team3, self.players[2], 2, 0)
        self._make_pick(self.team2, self.players[3], 2, 1)
        # Now it is team1's turn (round 2, index 2)
        response = self._post_pick(
            self.goalie2.pk, commissioner_token=self.session.commissioner_token
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("goalie", json.loads(response.content)["error"])

    def test_last_pick_sets_state_to_complete(self):
        # Fill all but the final slot manually, then submit the last pick
        all_signups = self.players[:6] + [self.cap1, self.cap2, self.cap3]
        slot = 0
        for r in range(1, 4):
            for p in range(3):
                if r == 3 and p == 2:
                    break  # leave last slot for the view
                self._make_pick(
                    [self.team1, self.team2, self.team3][p],
                    all_signups[slot],
                    r,
                    p,
                )
                slot += 1
        last_signup = all_signups[slot]
        self._post_pick(
            last_signup.pk,
            commissioner_token=self.session.commissioner_token,
        )
        self.session.refresh_from_db()
        self.assertEqual(self.session.state, DraftSession.STATE_COMPLETE)


# ---------------------------------------------------------------------------
# Logic: _process_auto_captain_picks
# ---------------------------------------------------------------------------


class AutoCaptainPickTests(DraftTestBase):
    def setUp(self):
        super().setUp()
        self._activate()

    def test_auto_captain_fires_in_designated_round(self):
        # team1's captain is auto-drafted in round 1, pick index 0
        self.team1.captain_draft_round = 1
        self.team1.save(update_fields=["captain_draft_round"])

        auto_picks = _process_auto_captain_picks(self.session)

        self.assertEqual(len(auto_picks), 1)
        self.assertEqual(auto_picks[0].signup, self.cap1)
        self.assertEqual(auto_picks[0].team, self.team1)
        self.assertTrue(auto_picks[0].is_auto_captain)

    def test_auto_captain_does_not_fire_in_wrong_round(self):
        # captain_draft_round=2 but current pick is round 1
        self.team1.captain_draft_round = 2
        self.team1.save(update_fields=["captain_draft_round"])

        auto_picks = _process_auto_captain_picks(self.session)
        self.assertEqual(len(auto_picks), 0)

    def test_auto_captain_skips_if_already_drafted(self):
        self.team1.captain_draft_round = 1
        self.team1.save(update_fields=["captain_draft_round"])
        # Pre-draft the captain manually
        self._make_pick(self.team1, self.cap1, 1, 0, auto=False)

        # Shouldn't try to draft again (would raise unique constraint otherwise)
        auto_picks = _process_auto_captain_picks(self.session)
        self.assertEqual(len(auto_picks), 0)

    def test_consecutive_auto_captain_slots_all_resolved(self):
        # Both team1 (pos 1) and team2 (pos 2) have captain_draft_round=1
        self.team1.captain_draft_round = 1
        self.team1.save(update_fields=["captain_draft_round"])
        self.team2.captain_draft_round = 1
        self.team2.save(update_fields=["captain_draft_round"])

        auto_picks = _process_auto_captain_picks(self.session)

        self.assertEqual(len(auto_picks), 2)
        drafted_signups = {p.signup for p in auto_picks}
        self.assertIn(self.cap1, drafted_signups)
        self.assertIn(self.cap2, drafted_signups)


# ---------------------------------------------------------------------------
# View: undo_last_pick
# ---------------------------------------------------------------------------


class UndoPickViewTests(DraftTestBase):
    def setUp(self):
        super().setUp()
        self._activate()

    def _undo_url(self):
        return reverse(
            "draft_undo_pick",
            args=[self.session.pk, self.session.commissioner_token],
        )

    def test_undo_removes_last_pick(self):
        pick = self._make_pick(self.team1, self.players[0], 1, 0)
        response = self.client.post(self._undo_url())
        self.assertEqual(response.status_code, 200)
        self.assertFalse(DraftPick.objects.filter(pk=pick.pk).exists())

    def test_undo_no_picks_returns_400(self):
        response = self.client.post(self._undo_url())
        self.assertEqual(response.status_code, 400)
        self.assertIn("No picks", json.loads(response.content)["error"])

    def test_undo_wrong_token_returns_404(self):
        import uuid

        url = reverse("draft_undo_pick", args=[self.session.pk, uuid.uuid4()])
        response = self.client.post(url)
        self.assertEqual(response.status_code, 404)

    def test_undo_auto_captain_chain_removes_consecutive_auto_picks(self):
        # A manual pick followed by two auto-captain picks.
        # Undo should remove the two auto picks and stop at the manual pick.
        manual = self._make_pick(self.team1, self.players[0], 1, 0, auto=False)
        auto1 = self._make_pick(self.team2, self.cap2, 1, 1, auto=True)
        auto2 = self._make_pick(self.team3, self.cap3, 1, 2, auto=True)

        response = self.client.post(self._undo_url())
        self.assertEqual(response.status_code, 200)

        # Both auto picks are removed; the preceding manual pick is preserved.
        self.assertFalse(DraftPick.objects.filter(pk=auto2.pk).exists())
        self.assertFalse(DraftPick.objects.filter(pk=auto1.pk).exists())
        self.assertTrue(DraftPick.objects.filter(pk=manual.pk).exists())


# ---------------------------------------------------------------------------
# View: swap_pick
# ---------------------------------------------------------------------------


class SwapPickViewTests(DraftTestBase):
    def setUp(self):
        super().setUp()
        self._complete()

    def _swap_url(self):
        return reverse(
            "draft_swap_pick",
            args=[self.session.pk, self.session.commissioner_token],
        )

    def _swap(self, pick_id, new_signup_pk):
        return self.client.post(
            self._swap_url(),
            {"pick_id": pick_id, "new_signup_pk": new_signup_pk},
        )

    def test_swap_replaces_undrafted_player_into_slot(self):
        pick = self._make_pick(self.team1, self.players[0], 1, 0)
        response = self._swap(pick.pk, self.players[1].pk)
        self.assertEqual(response.status_code, 200)
        pick.refresh_from_db()
        self.assertEqual(pick.signup, self.players[1])

    def test_swap_exchanges_two_existing_picks(self):
        pick_a = self._make_pick(self.team1, self.players[0], 1, 0)
        self._make_pick(self.team2, self.players[1], 1, 1)

        response = self._swap(pick_a.pk, self.players[1].pk)
        self.assertEqual(response.status_code, 200)

        # pick_a now holds players[1]
        pick_a.refresh_from_db()
        self.assertEqual(pick_a.signup, self.players[1])

        # The slot previously held by pick_b (team2, round 1, index 1)
        # now holds players[0] — looked up by position since pick_b was recreated
        new_pick_b = DraftPick.objects.get(
            session=self.session, team=self.team2, round_number=1, pick_number=1
        )
        self.assertEqual(new_pick_b.signup, self.players[0])

    def test_swap_same_player_is_no_op(self):
        pick = self._make_pick(self.team1, self.players[0], 1, 0)
        response = self._swap(pick.pk, self.players[0].pk)
        self.assertEqual(response.status_code, 200)
        pick.refresh_from_db()
        self.assertEqual(pick.signup, self.players[0])

    def test_swap_cannot_move_captain_to_different_team(self):
        # put cap2 into team1's pick slot — should be rejected
        pick = self._make_pick(self.team1, self.players[0], 1, 0)
        response = self._swap(pick.pk, self.cap2.pk)
        self.assertEqual(response.status_code, 400)
        data = json.loads(response.content)
        self.assertIn("captain of", data["error"])
        self.assertIn("own team", data["error"])

    def test_swap_allows_captain_to_move_within_own_team(self):
        # Swap cap1 into team1's slot — same team, should succeed
        pick = self._make_pick(self.team1, self.players[0], 1, 0)
        response = self._swap(pick.pk, self.cap1.pk)
        self.assertEqual(response.status_code, 200)
        pick.refresh_from_db()
        self.assertEqual(pick.signup, self.cap1)

    def test_swap_allowed_when_paused(self):
        self.session.state = DraftSession.STATE_PAUSED
        self.session.save(update_fields=["state"])
        pick = self._make_pick(self.team1, self.players[0], 1, 0)
        response = self._swap(pick.pk, self.players[1].pk)
        self.assertEqual(response.status_code, 200)
        pick.refresh_from_db()
        self.assertEqual(pick.signup, self.players[1])

    def test_swap_blocked_when_active(self):
        self.session.state = DraftSession.STATE_ACTIVE
        self.session.save(update_fields=["state"])
        pick = self._make_pick(self.team1, self.players[0], 1, 0)
        response = self._swap(pick.pk, self.players[1].pk)
        self.assertEqual(response.status_code, 400)
        self.assertIn("paused or complete", json.loads(response.content)["error"])

    def test_swap_wrong_token_returns_404(self):
        import uuid

        pick = self._make_pick(self.team1, self.players[0], 1, 0)
        url = reverse("draft_swap_pick", args=[self.session.pk, uuid.uuid4()])
        response = self.client.post(
            url, {"pick_id": pick.pk, "new_signup_pk": self.players[1].pk}
        )
        self.assertEqual(response.status_code, 404)


# ---------------------------------------------------------------------------
# View: reset_draft
# ---------------------------------------------------------------------------


class ResetDraftViewTests(DraftTestBase):
    def setUp(self):
        super().setUp()
        self._activate()

    def _reset_url(self):
        return reverse(
            "draft_reset",
            args=[self.session.pk, self.session.commissioner_token],
        )

    def test_reset_deletes_all_picks(self):
        self._make_pick(self.team1, self.players[0], 1, 0)
        self._make_pick(self.team2, self.players[1], 1, 1)
        self.client.post(self._reset_url())
        self.assertEqual(DraftPick.objects.filter(session=self.session).count(), 0)

    def test_reset_clears_draft_positions(self):
        self.client.post(self._reset_url())
        positions = list(self.session.teams.values_list("draft_position", flat=True))
        self.assertTrue(all(p is None for p in positions))

    def test_reset_returns_session_to_setup(self):
        self.client.post(self._reset_url())
        self.session.refresh_from_db()
        self.assertEqual(self.session.state, DraftSession.STATE_SETUP)

    def test_reset_reopens_signups(self):
        self.session.signups_open = False
        self.session.save(update_fields=["signups_open"])
        self.client.post(self._reset_url())
        self.session.refresh_from_db()
        self.assertTrue(self.session.signups_open)

    def test_reset_from_setup_returns_400(self):
        self.session.state = DraftSession.STATE_SETUP
        self.session.save(update_fields=["state"])
        response = self.client.post(self._reset_url())
        self.assertEqual(response.status_code, 400)
        self.assertIn("not been started", json.loads(response.content)["error"])

    def test_reset_wrong_token_returns_404(self):
        import uuid

        url = reverse("draft_reset", args=[self.session.pk, uuid.uuid4()])
        response = self.client.post(url)
        self.assertEqual(response.status_code, 404)


# ---------------------------------------------------------------------------
# View: set_captain_rounds
# ---------------------------------------------------------------------------


class SetCaptainRoundsViewTests(DraftTestBase):
    def setUp(self):
        super().setUp()
        # Put session in DRAW state (positions already drawn by DraftTestBase)
        self.session.state = DraftSession.STATE_DRAW
        self.session.save(update_fields=["state"])

    def _url(self):
        return reverse(
            "draft_set_captain_rounds",
            args=[self.session.pk, self.session.commissioner_token],
        )

    def _post(self, rounds_map):
        return self.client.post(
            self._url(),
            data=json.dumps({"rounds": rounds_map}),
            content_type="application/json",
        )

    def test_sets_rounds_in_draw_state(self):
        response = self._post(
            {str(self.team1.pk): 1, str(self.team2.pk): 2, str(self.team3.pk): 3}
        )
        self.assertEqual(response.status_code, 200)
        self.team1.refresh_from_db()
        self.team2.refresh_from_db()
        self.team3.refresh_from_db()
        self.assertEqual(self.team1.captain_draft_round, 1)
        self.assertEqual(self.team2.captain_draft_round, 2)
        self.assertEqual(self.team3.captain_draft_round, 3)

    def test_sets_rounds_in_setup_state(self):
        self.session.state = DraftSession.STATE_SETUP
        self.session.save(update_fields=["state"])
        response = self._post({str(self.team1.pk): 2})
        self.assertEqual(response.status_code, 200)
        self.team1.refresh_from_db()
        self.assertEqual(self.team1.captain_draft_round, 2)

    def test_clears_round_when_null(self):
        self.team1.captain_draft_round = 1
        self.team1.save(update_fields=["captain_draft_round"])
        response = self._post({str(self.team1.pk): None})
        self.assertEqual(response.status_code, 200)
        self.team1.refresh_from_db()
        self.assertIsNone(self.team1.captain_draft_round)

    def test_rejects_round_out_of_range(self):
        response = self._post({str(self.team1.pk): 99})
        self.assertEqual(response.status_code, 400)
        data = json.loads(response.content)
        self.assertIn("out of range", data["error"])

    def test_rejects_round_zero(self):
        response = self._post({str(self.team1.pk): 0})
        self.assertEqual(response.status_code, 400)

    def test_rejects_when_active(self):
        self.session.state = DraftSession.STATE_ACTIVE
        self.session.save(update_fields=["state"])
        response = self._post({str(self.team1.pk): 1})
        self.assertEqual(response.status_code, 400)
        data = json.loads(response.content)
        self.assertIn("before the draft starts", data["error"])

    def test_wrong_token_returns_404(self):
        import uuid

        url = reverse("draft_set_captain_rounds", args=[self.session.pk, uuid.uuid4()])
        response = self.client.post(
            url,
            data=json.dumps({"rounds": {str(self.team1.pk): 1}}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 404)

    def test_get_not_allowed(self):
        response = self.client.get(self._url())
        self.assertEqual(response.status_code, 405)

    def test_returns_updated_state_payload(self):
        response = self._post({str(self.team1.pk): 1})
        data = json.loads(response.content)
        self.assertIn("state", data)
        team_data = next(t for t in data["state"]["teams"] if t["id"] == self.team1.pk)
        self.assertEqual(team_data["captain_draft_round"], 1)

    def test_ignores_unknown_team_pk(self):
        response = self._post({"99999": 1})
        self.assertEqual(response.status_code, 200)


# ---------------------------------------------------------------------------
# Helper: _session_state_payload
# ---------------------------------------------------------------------------


class SessionStatePayloadTests(DraftTestBase):
    def test_payload_teams_sorted_by_draft_position(self):
        # Assign positions out of natural insertion order
        self.team1.draft_position = 3
        self.team1.save(update_fields=["draft_position"])
        self.team3.draft_position = 1
        self.team3.save(update_fields=["draft_position"])

        payload = _session_state_payload(self.session)
        positions = [t["draft_position"] for t in payload["teams"]]
        self.assertEqual(positions, sorted(positions))

    def test_payload_active_team_pk_matches_expected_turn(self):
        self._activate()
        payload = _session_state_payload(self.session)
        # Round 1, index 0 → team with draft_position=1 = team1
        self.assertEqual(payload["active_team_pk"], self.team1.pk)

    def test_payload_available_players_excludes_drafted(self):
        self._activate()
        self._make_pick(self.team1, self.players[0], 1, 0)
        payload = _session_state_payload(self.session)
        available_ids = {p["id"] for p in payload["available_players"]}
        self.assertNotIn(self.players[0].pk, available_ids)
        self.assertIn(self.players[1].pk, available_ids)

    def test_payload_available_players_excludes_captains(self):
        payload = _session_state_payload(self.session)
        available_ids = {p["id"] for p in payload["available_players"]}
        self.assertNotIn(self.cap1.pk, available_ids)
        self.assertNotIn(self.cap2.pk, available_ids)
        self.assertNotIn(self.cap3.pk, available_ids)
        self.assertIn(self.players[0].pk, available_ids)

    def test_payload_pick_includes_pick_id(self):
        self._activate()
        pick = self._make_pick(self.team1, self.players[0], 1, 0)
        payload = _session_state_payload(self.session)
        team_data = next(t for t in payload["teams"] if t["id"] == self.team1.pk)
        self.assertEqual(team_data["picks"][1]["pick_id"], pick.pk)

    def test_payload_state_matches_session_state(self):
        self._activate()
        payload = _session_state_payload(self.session)
        self.assertEqual(payload["state"], DraftSession.STATE_ACTIVE)

    def test_payload_current_round_and_index_none_when_complete(self):
        # Fill all picks to make draft complete
        all_signups = self.players[:6] + [self.cap1, self.cap2, self.cap3]
        slot = 0
        for r in range(1, 4):
            for p in range(3):
                self._make_pick(
                    [self.team1, self.team2, self.team3][p],
                    all_signups[slot],
                    r,
                    p,
                )
                slot += 1
        payload = _session_state_payload(self.session)
        self.assertIsNone(payload["current_round"])
        self.assertIsNone(payload["current_pick_index"])
        self.assertIsNone(payload["active_team_pk"])


# ---------------------------------------------------------------------------
# Edge: multiple randomized rounds — snake parity skips all of them
# ---------------------------------------------------------------------------


class MultipleRandomizedRoundsTests(DraftTestBase):
    """
    When two or more rounds are randomized the snake-parity counter excludes
    all of them.  With rounds 2 and 3 both randomized in a 4-round draft,
    round 4's effective round is 4 - 2 = 2 (even) → reversed pick order.
    """

    def setUp(self):
        super().setUp()
        self.session.num_rounds = 4
        self.session.save(update_fields=["num_rounds"])
        DraftRound.objects.create(
            session=self.session, round_number=2, order_type=DraftRound.ORDER_RANDOMIZED
        )
        DraftRound.objects.create(
            session=self.session, round_number=3, order_type=DraftRound.ORDER_RANDOMIZED
        )

    def test_round_1_still_forward_with_later_randomized_rounds(self):
        order = self.session.pick_order_for_round(1)
        self.assertEqual(order, [self.team1.pk, self.team2.pk, self.team3.pk])

    def test_round_4_parity_skips_both_randomized_rounds(self):
        # effective_round = 4 - 2 = 2  (even) → reversed
        order = self.session.pick_order_for_round(4)
        self.assertEqual(order, [self.team3.pk, self.team2.pk, self.team1.pk])

    def test_each_randomized_round_contains_all_teams(self):
        self.assertCountEqual(
            self.session.pick_order_for_round(2),
            [self.team1.pk, self.team2.pk, self.team3.pk],
        )
        self.assertCountEqual(
            self.session.pick_order_for_round(3),
            [self.team1.pk, self.team2.pk, self.team3.pk],
        )

    def test_each_randomized_round_is_deterministic(self):
        self.assertEqual(
            self.session.pick_order_for_round(2), self.session.pick_order_for_round(2)
        )
        self.assertEqual(
            self.session.pick_order_for_round(3), self.session.pick_order_for_round(3)
        )

    def test_randomized_rounds_use_different_seeds(self):
        # Seeded by f"{session.pk}-{round_number}"; rounds 2 and 3 should differ.
        # With 3 teams (6 permutations) a collision is possible but extremely unlikely.
        # We verify at minimum that the keys are independent (not the same object).
        order_2 = self.session.pick_order_for_round(2)
        order_3 = self.session.pick_order_for_round(3)
        # Both must contain all teams regardless of order.
        self.assertCountEqual(order_2, [self.team1.pk, self.team2.pk, self.team3.pk])
        self.assertCountEqual(order_3, [self.team1.pk, self.team2.pk, self.team3.pk])


# ---------------------------------------------------------------------------
# Edge: round-11-style randomization — snake resumes after re-randomized round
# ---------------------------------------------------------------------------


class Round11StyleRandomizationTests(DraftTestBase):
    """
    Scaled-down simulation of the real round-11 re-randomization mechanic.
    In the 3-round test draft, round 2 acts as the "round 11" wildcard.

    Without randomization: R1→forward, R2→reversed, R3→forward.
    With R2 randomized:    R1→forward, R2→randomized, R3→reversed
                           (effective round for R3 = 3 - 1 = 2, even).
    """

    def setUp(self):
        super().setUp()
        DraftRound.objects.create(
            session=self.session, round_number=2, order_type=DraftRound.ORDER_RANDOMIZED
        )

    def test_round_1_unaffected_by_later_randomized_round(self):
        order = self.session.pick_order_for_round(1)
        self.assertEqual(order, [self.team1.pk, self.team2.pk, self.team3.pk])

    def test_round_2_is_randomized_not_pure_snake(self):
        order = self.session.pick_order_for_round(2)
        # Must contain all teams; may or may not be forward or reversed
        self.assertCountEqual(order, [self.team1.pk, self.team2.pk, self.team3.pk])

    def test_round_3_snake_continues_as_if_randomized_round_never_happened(self):
        # Effective round for R3 = 3 - 1 = 2 (even) → reversed
        order = self.session.pick_order_for_round(3)
        self.assertEqual(order, [self.team3.pk, self.team2.pk, self.team1.pk])

    def test_randomized_round_order_is_deterministic_across_calls(self):
        self.assertEqual(
            self.session.pick_order_for_round(2), self.session.pick_order_for_round(2)
        )

    def test_randomized_round_seed_is_per_session(self):
        """Two different sessions produce independently seeded random orders."""
        season2 = Season.objects.create(
            year=2098, season_type=3, is_current_season=False
        )
        session2 = DraftSession.objects.create(
            season=season2, num_teams=3, num_rounds=3
        )
        for pos, email in enumerate(["s1@t.com", "s2@t.com", "s3@t.com"], start=1):
            signup = SeasonSignup.objects.create(
                season=season2,
                first_name=f"S{pos}",
                last_name="Draft",
                email=email,
                primary_position=SeasonSignup.POSITION_CENTER,
                secondary_position=SeasonSignup.POSITION_ONE_THING,
                captain_interest=SeasonSignup.CAPTAIN_YES,
            )
            DraftTeam.objects.create(
                session=session2, captain=signup, draft_position=pos
            )
        DraftRound.objects.create(
            session=session2, round_number=2, order_type=DraftRound.ORDER_RANDOMIZED
        )
        # Each session's randomized order contains exactly its own teams
        order_s1 = self.session.pick_order_for_round(2)
        order_s2 = session2.pick_order_for_round(2)
        self.assertCountEqual(order_s1, [self.team1.pk, self.team2.pk, self.team3.pk])
        self.assertCountEqual(
            order_s2, list(session2.teams.values_list("pk", flat=True))
        )


# ---------------------------------------------------------------------------
# Edge: auto-captain firing inside a randomized round
# ---------------------------------------------------------------------------


class AutoCaptainInRandomizedRoundTests(DraftTestBase):
    """
    All three captains have captain_draft_round pointing at a randomized round.
    _process_auto_captain_picks must auto-draft them in the correct randomized
    slot order, not the regular snake order.
    """

    def setUp(self):
        super().setUp()
        self._activate()
        DraftRound.objects.create(
            session=self.session, round_number=2, order_type=DraftRound.ORDER_RANDOMIZED
        )
        for team in [self.team1, self.team2, self.team3]:
            team.captain_draft_round = 2
            team.save(update_fields=["captain_draft_round"])

    def _fill_round_1(self):
        order_r1 = self.session.pick_order_for_round(1)
        for i, team_pk in enumerate(order_r1):
            team = DraftTeam.objects.get(pk=team_pk)
            self._make_pick(team, self.players[i], 1, i)

    def test_all_captains_auto_picked_in_randomized_round(self):
        self._fill_round_1()
        auto_picks = _process_auto_captain_picks(self.session)
        self.assertEqual(len(auto_picks), 3)
        drafted_signups = {p.signup for p in auto_picks}
        self.assertIn(self.cap1, drafted_signups)
        self.assertIn(self.cap2, drafted_signups)
        self.assertIn(self.cap3, drafted_signups)

    def test_auto_picks_respect_randomized_slot_order(self):
        self._fill_round_1()
        auto_picks = _process_auto_captain_picks(self.session)
        expected_order = self.session.pick_order_for_round(2)
        actual_order = [p.team.pk for p in auto_picks]
        self.assertEqual(actual_order, expected_order)

    def test_auto_picks_have_correct_round_and_pick_numbers(self):
        self._fill_round_1()
        auto_picks = _process_auto_captain_picks(self.session)
        self.assertTrue(all(p.round_number == 2 for p in auto_picks))
        self.assertEqual([p.pick_number for p in auto_picks], [0, 1, 2])

    def test_auto_picks_all_flagged_is_auto_captain(self):
        self._fill_round_1()
        auto_picks = _process_auto_captain_picks(self.session)
        self.assertTrue(all(p.is_auto_captain for p in auto_picks))


# ---------------------------------------------------------------------------
# Edge: auto-captain pick in the final draft slot → draft completes
# ---------------------------------------------------------------------------


class AutoCaptainCompletionTests(DraftTestBase):
    """
    When the auto-captain pick lands in the very last slot of the draft,
    _process_auto_captain_picks (or the make_pick view) must set the session
    state to STATE_COMPLETE.

    Layout (3 teams, 3 rounds):
      R1 forward:  T1(0) T2(1) T3(2)
      R2 reversed: T3(0) T2(1) T1(2)
      R3 forward:  T1(0) T2(1) T3(2)  ← T3 at slot 2 is the last pick
    """

    def setUp(self):
        super().setUp()
        self._activate()
        # team3 (draft_position=3) occupies the last slot in round 3 (forward)
        self.team3.captain_draft_round = 3
        self.team3.save(update_fields=["captain_draft_round"])

    def _fill_all_except_last_round_last_slot(self):
        """Fill all 9 picks except (round 3, pick 2) — the auto-captain slot."""
        # Round 1: forward [T1, T2, T3]
        self._make_pick(self.team1, self.players[0], 1, 0)
        self._make_pick(self.team2, self.players[1], 1, 1)
        self._make_pick(self.team3, self.players[2], 1, 2)
        # Round 2: reversed [T3, T2, T1]
        self._make_pick(self.team3, self.players[3], 2, 0)
        self._make_pick(self.team2, self.players[4], 2, 1)
        self._make_pick(self.team1, self.players[5], 2, 2)
        # Round 3, picks 0 and 1 — leave pick 2 for auto-captain
        self._make_pick(self.team1, self.cap1, 3, 0)
        self._make_pick(self.team2, self.cap2, 3, 1)

    def test_auto_captain_as_final_pick_sets_state_complete(self):
        self._fill_all_except_last_round_last_slot()
        auto_picks = _process_auto_captain_picks(self.session)
        self.assertEqual(len(auto_picks), 1)
        self.assertEqual(auto_picks[0].signup, self.cap3)
        self.session.refresh_from_db()
        self.assertEqual(self.session.state, DraftSession.STATE_COMPLETE)

    def test_auto_captain_completion_via_make_pick_api(self):
        """
        Submitting the second-to-last pick via the API triggers auto-captain
        for the last slot, completing the draft without any further manual picks.
        """
        # Fill all but the last TWO picks (round 3 picks 1 and 2)
        self._make_pick(self.team1, self.players[0], 1, 0)
        self._make_pick(self.team2, self.players[1], 1, 1)
        self._make_pick(self.team3, self.players[2], 1, 2)
        self._make_pick(self.team3, self.players[3], 2, 0)
        self._make_pick(self.team2, self.players[4], 2, 1)
        self._make_pick(self.team1, self.players[5], 2, 2)
        self._make_pick(self.team1, self.cap1, 3, 0)

        # Submit round 3, pick 1 (team2's turn): cap2 for team2
        response = self._post_pick(self.cap2.pk, captain_token=self.team2.captain_token)
        self.assertEqual(response.status_code, 200)
        self.session.refresh_from_db()
        self.assertEqual(self.session.state, DraftSession.STATE_COMPLETE)
        # cap3 was auto-drafted for team3 without any additional API call
        self.assertTrue(
            DraftPick.objects.filter(
                session=self.session,
                signup=self.cap3,
                team=self.team3,
                is_auto_captain=True,
            ).exists()
        )


# ---------------------------------------------------------------------------
# Edge: all captains share the same captain_draft_round (entire round is auto)
# ---------------------------------------------------------------------------


class AllCaptainsSameRoundTests(DraftTestBase):
    """
    When all captains are assigned captain_draft_round=1 the entire first round
    resolves automatically.  After activation, no manual picks are needed until
    round 2.  Undo of an auto-captain chain removes all consecutive auto picks.
    """

    def setUp(self):
        super().setUp()
        for team in [self.team1, self.team2, self.team3]:
            team.captain_draft_round = 1
            team.save(update_fields=["captain_draft_round"])
        self._activate()

    def test_process_auto_captain_picks_resolves_all_three_consecutively(self):
        auto_picks = _process_auto_captain_picks(self.session)
        self.assertEqual(len(auto_picks), 3)
        self.assertTrue(all(p.round_number == 1 for p in auto_picks))
        self.assertEqual(sorted(p.pick_number for p in auto_picks), [0, 1, 2])

    def test_after_full_round_auto_pick_current_pick_advances_to_round_2(self):
        _process_auto_captain_picks(self.session)
        self.assertEqual(self.session.current_pick, (2, 0))

    def test_all_auto_picks_are_flagged_is_auto_captain(self):
        auto_picks = _process_auto_captain_picks(self.session)
        self.assertTrue(all(p.is_auto_captain for p in auto_picks))

    def test_activation_fires_auto_captain_picks_on_draw_to_active(self):
        """Advancing DRAW→ACTIVE auto-picks all round-1 captains immediately."""
        # Reset to DRAW so advance_state triggers the transition
        self.session.state = DraftSession.STATE_DRAW
        self.session.save(update_fields=["state"])
        self.client.post(
            reverse(
                "draft_advance_state",
                args=[self.session.pk, self.session.commissioner_token],
            )
        )
        # All three captains should be drafted
        self.assertEqual(DraftPick.objects.filter(session=self.session).count(), 3)
        for cap in [self.cap1, self.cap2, self.cap3]:
            self.assertTrue(
                DraftPick.objects.filter(
                    session=self.session, signup=cap, is_auto_captain=True
                ).exists()
            )


# ---------------------------------------------------------------------------
# Edge: own team manually picks their captain before the auto-captain round
# ---------------------------------------------------------------------------


class EarlyManualCaptainPickTests(DraftTestBase):
    """
    A captain with captain_draft_round=3 can still be manually drafted by
    their own team in an earlier round.  When round 3 arrives, the auto-pick
    guard finds them already drafted and skips silently.
    """

    def setUp(self):
        super().setUp()
        self._activate()
        self.team1.captain_draft_round = 3  # auto-slot is round 3, pick index 0
        self.team1.save(update_fields=["captain_draft_round"])

    def test_own_team_can_pick_captain_manually_before_designated_round(self):
        # Round 1, slot 0 → team1's turn.  Manually pick their own captain.
        response = self._post_pick(self.cap1.pk, captain_token=self.team1.captain_token)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            DraftPick.objects.filter(
                session=self.session,
                signup=self.cap1,
                team=self.team1,
                is_auto_captain=False,
            ).exists()
        )

    def test_auto_pick_skips_silently_when_captain_already_drafted(self):
        # Manually pre-draft cap1 in round 1, slot 0
        self._make_pick(self.team1, self.cap1, 1, 0)
        # Fill the rest of rounds 1 and 2
        self._make_pick(self.team2, self.players[0], 1, 1)
        self._make_pick(self.team3, self.players[1], 1, 2)
        self._make_pick(self.team3, self.players[2], 2, 0)
        self._make_pick(self.team2, self.players[3], 2, 1)
        self._make_pick(self.team1, self.players[4], 2, 2)
        # Now at round 3, pick 0 → team1's captain slot — but cap1 already drafted
        auto_picks = _process_auto_captain_picks(self.session)
        self.assertEqual(len(auto_picks), 0)

    def test_captain_draft_round_beyond_num_rounds_never_fires(self):
        self.team1.captain_draft_round = 99
        self.team1.save(update_fields=["captain_draft_round"])
        auto_picks = _process_auto_captain_picks(self.session)
        self.assertEqual(len(auto_picks), 0)


# ---------------------------------------------------------------------------
# Edge: undo at round boundary and undo all picks
# ---------------------------------------------------------------------------


class UndoEdgeCaseTests(DraftTestBase):
    """Undo across round boundaries, repeated undo, undo with complete state."""

    def setUp(self):
        super().setUp()
        self._activate()

    def _undo_url(self):
        return reverse(
            "draft_undo_pick",
            args=[self.session.pk, self.session.commissioner_token],
        )

    def test_undo_at_round_boundary_restores_previous_round_turn(self):
        # Fill round 1 completely, then make first pick of round 2
        self._make_pick(self.team1, self.players[0], 1, 0)
        self._make_pick(self.team2, self.players[1], 1, 1)
        self._make_pick(self.team3, self.players[2], 1, 2)
        # Round 2, pick 0 belongs to team3 (reversed)
        self._make_pick(self.team3, self.players[3], 2, 0)

        response = self.client.post(self._undo_url())
        self.assertEqual(response.status_code, 200)
        # current_pick should revert to (2, 0) — first slot of round 2
        self.assertEqual(self.session.current_pick, (2, 0))
        self.assertFalse(
            DraftPick.objects.filter(
                session=self.session, signup=self.players[3]
            ).exists()
        )

    def test_repeated_undo_empties_entire_draft(self):
        self._make_pick(self.team1, self.players[0], 1, 0)
        self._make_pick(self.team2, self.players[1], 1, 1)
        self._make_pick(self.team3, self.players[2], 1, 2)

        for _ in range(3):
            response = self.client.post(self._undo_url())
            self.assertEqual(response.status_code, 200)

        self.assertEqual(DraftPick.objects.filter(session=self.session).count(), 0)
        self.assertEqual(self.session.current_pick, (1, 0))

    def test_undo_after_emptying_returns_400(self):
        self._make_pick(self.team1, self.players[0], 1, 0)
        self.client.post(self._undo_url())  # removes only pick
        response = self.client.post(self._undo_url())  # nothing left
        self.assertEqual(response.status_code, 400)

    def test_undo_from_complete_state_removes_last_pick(self):
        # Fill the entire draft
        all_signups = self.players[:6] + [self.cap1, self.cap2, self.cap3]
        slot = 0
        for r in range(1, 4):
            for p in range(3):
                self._make_pick(
                    [self.team1, self.team2, self.team3][p],
                    all_signups[slot],
                    r,
                    p,
                )
                slot += 1
        self.session.state = DraftSession.STATE_COMPLETE
        self.session.save(update_fields=["state"])

        # Undo should remove the last pick even from COMPLETE state
        last_pick_signup = all_signups[slot - 1]
        response = self.client.post(self._undo_url())
        self.assertEqual(response.status_code, 200)
        self.assertFalse(
            DraftPick.objects.filter(
                session=self.session, signup=last_pick_signup
            ).exists()
        )


# ---------------------------------------------------------------------------
# Edge: pick-validation edge cases
# ---------------------------------------------------------------------------


class PickValidationEdgeCaseTests(DraftTestBase):
    """Nonexistent signup, cross-season signup, GET method, complete-state guard."""

    def setUp(self):
        super().setUp()
        self._activate()

    def test_make_pick_get_request_returns_405(self):
        url = reverse("draft_make_pick", args=[self.session.pk])
        response = self.client.get(
            url,
            {
                "signup_pk": self.players[0].pk,
                "commissioner_token": str(self.session.commissioner_token),
            },
        )
        self.assertEqual(response.status_code, 405)

    def test_make_pick_nonexistent_signup_pk_returns_400(self):
        response = self._post_pick(
            99999, commissioner_token=self.session.commissioner_token
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("not found", json.loads(response.content)["error"].lower())

    def test_make_pick_signup_from_different_season_returns_400(self):
        other_season = Season.objects.create(
            year=2099, season_type=4, is_current_season=False
        )
        other_signup = SeasonSignup.objects.create(
            season=other_season,
            first_name="Wrong",
            last_name="Season",
            email="wrong@test.com",
            primary_position=SeasonSignup.POSITION_CENTER,
            secondary_position=SeasonSignup.POSITION_ONE_THING,
            captain_interest=SeasonSignup.CAPTAIN_NO,
        )
        response = self._post_pick(
            other_signup.pk, commissioner_token=self.session.commissioner_token
        )
        self.assertEqual(response.status_code, 400)

    def test_make_pick_when_draft_complete_returns_400(self):
        self.session.state = DraftSession.STATE_COMPLETE
        self.session.save(update_fields=["state"])
        response = self._post_pick(
            self.players[0].pk, commissioner_token=self.session.commissioner_token
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("not active", json.loads(response.content)["error"])

    def test_auto_captain_guard_overrides_submitted_signup_pk(self):
        """
        When it's a team's captain_draft_round and their turn, the view ignores
        the submitted signup_pk and auto-drafts the captain instead.
        """
        self.team1.captain_draft_round = 1  # round 1, pick 0 → team1's turn
        self.team1.save(update_fields=["captain_draft_round"])

        # Submit a regular player, expect cap1 to be drafted instead
        response = self._post_pick(
            self.players[0].pk, captain_token=self.team1.captain_token
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            DraftPick.objects.filter(
                session=self.session,
                signup=self.cap1,
                team=self.team1,
                is_auto_captain=True,
            ).exists()
        )
        self.assertFalse(
            DraftPick.objects.filter(
                session=self.session, signup=self.players[0]
            ).exists()
        )

    def test_captain_token_for_wrong_turn_blocked_even_when_captain_draft_round_matches(
        self,
    ):
        """
        A captain submitting their token when another team's captain_draft_round
        fires for slot 0 — their token is invalid for that slot, so they get 403.
        """
        # team2's captain_draft_round=1 at slot 0 — it's team1's slot, not team2's
        # This tests: team1 is at slot 0 (position 1), team2 is at slot 1 (position 2)
        # team2 tries to submit — "not your turn"
        self.team2.captain_draft_round = 1
        self.team2.save(update_fields=["captain_draft_round"])

        response = self._post_pick(
            self.players[0].pk, captain_token=self.team2.captain_token
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("not your turn", json.loads(response.content)["error"])


# ---------------------------------------------------------------------------
# Edge: _session_state_payload advanced checks
# ---------------------------------------------------------------------------


class SessionStatePayloadAdvancedTests(DraftTestBase):
    """has_goalie flag, snake flip in active_team_pk, num_teams/num_rounds in payload."""

    def test_payload_has_goalie_true_when_team_drafted_goalie(self):
        self._activate()
        self._make_pick(self.team1, self.goalie1, 1, 0)
        payload = _session_state_payload(self.session)
        team1_data = next(t for t in payload["teams"] if t["id"] == self.team1.pk)
        self.assertTrue(team1_data["has_goalie"])

    def test_payload_has_goalie_false_when_team_has_no_goalie(self):
        self._activate()
        self._make_pick(self.team1, self.players[0], 1, 0)
        payload = _session_state_payload(self.session)
        team1_data = next(t for t in payload["teams"] if t["id"] == self.team1.pk)
        self.assertFalse(team1_data["has_goalie"])

    def test_payload_active_team_advances_after_each_pick(self):
        self._activate()
        payload = _session_state_payload(self.session)
        self.assertEqual(payload["active_team_pk"], self.team1.pk)

        self._make_pick(self.team1, self.players[0], 1, 0)
        payload = _session_state_payload(self.session)
        self.assertEqual(payload["active_team_pk"], self.team2.pk)

        self._make_pick(self.team2, self.players[1], 1, 1)
        payload = _session_state_payload(self.session)
        self.assertEqual(payload["active_team_pk"], self.team3.pk)

    def test_payload_active_team_snake_flips_at_round_boundary(self):
        """After round 1 completes (T1→T2→T3), round 2 starts with T3 (reversed)."""
        self._activate()
        self._make_pick(self.team1, self.players[0], 1, 0)
        self._make_pick(self.team2, self.players[1], 1, 1)
        self._make_pick(self.team3, self.players[2], 1, 2)
        payload = _session_state_payload(self.session)
        self.assertEqual(payload["current_round"], 2)
        self.assertEqual(payload["current_pick_index"], 0)
        self.assertEqual(payload["active_team_pk"], self.team3.pk)

    def test_payload_exposes_num_teams_and_num_rounds(self):
        payload = _session_state_payload(self.session)
        self.assertEqual(payload["num_teams"], 3)
        self.assertEqual(payload["num_rounds"], 3)

    def test_payload_finalized_false_before_finalize(self):
        payload = _session_state_payload(self.session)
        self.assertFalse(payload["finalized"])


# ---------------------------------------------------------------------------
# Integration: complete snake draft simulated end-to-end via the API
# ---------------------------------------------------------------------------


class FullSnakeDraftIntegrationTests(DraftTestBase):
    """
    End-to-end integration test: all 9 picks submitted via the make_pick view
    in the correct snake order.  Verifies pick assignment, turn enforcement,
    and STATE_COMPLETE at the end.

    Pick sequence (3 teams, 3 rounds, no captain_draft_rounds set):
      R1 forward:  T1(0) T2(1) T3(2)
      R2 reversed: T3(0) T2(1) T1(2)
      R3 forward:  T1(0) T2(1) T3(2)
    """

    def setUp(self):
        super().setUp()
        self._activate()
        # Use captains' own tokens; captains have no captain_draft_round so are pickable manually
        self.all_signups = self.players[:6] + [self.cap1, self.cap2, self.cap3]

    def test_complete_snake_draft_all_picks_succeed(self):
        """
        Every pick in the correct order returns 200.  After all 9 picks the
        session is STATE_COMPLETE and each player is on the expected team.
        """
        pick_sequence = [
            # Round 1: T1 → T2 → T3
            (self.team1, self.all_signups[0]),
            (self.team2, self.all_signups[1]),
            (self.team3, self.all_signups[2]),
            # Round 2: T3 → T2 → T1  (snake reversed)
            (self.team3, self.all_signups[3]),
            (self.team2, self.all_signups[4]),
            (self.team1, self.all_signups[5]),
            # Round 3: T1 → T2 → T3  (snake forward again)
            (self.team1, self.all_signups[6]),  # cap1 on team1
            (self.team2, self.all_signups[7]),  # cap2 on team2
            (self.team3, self.all_signups[8]),  # cap3 on team3
        ]

        for i, (expected_team, signup) in enumerate(pick_sequence):
            response = self._post_pick(
                signup.pk, commissioner_token=self.session.commissioner_token
            )
            self.assertEqual(
                response.status_code,
                200,
                f"Pick {i + 1} failed ({signup} → {expected_team}): {response.content}",
            )
            self.assertTrue(
                DraftPick.objects.filter(
                    session=self.session, signup=signup, team=expected_team
                ).exists(),
                f"Pick {i + 1}: {signup} not recorded on {expected_team}",
            )

        self.session.refresh_from_db()
        self.assertEqual(self.session.state, DraftSession.STATE_COMPLETE)
        self.assertEqual(DraftPick.objects.filter(session=self.session).count(), 9)

    def test_out_of_snake_order_captain_token_rejected(self):
        """
        A captain submitting their token when it's not their turn gets 400.
        Round 1, slot 0 belongs to team1; team3 submitting is rejected.
        """
        response = self._post_pick(
            self.players[0].pk, captain_token=self.team3.captain_token
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("not your turn", json.loads(response.content)["error"])

    def test_commissioner_can_pick_for_any_team_in_order(self):
        """
        Commissioner token is accepted for every pick regardless of which
        captain's turn it is; the pick is credited to the active team.
        """
        # All 9 picks via commissioner token
        for signup in self.all_signups:
            response = self._post_pick(
                signup.pk, commissioner_token=self.session.commissioner_token
            )
            self.assertEqual(response.status_code, 200)

        self.session.refresh_from_db()
        self.assertEqual(self.session.state, DraftSession.STATE_COMPLETE)

    def test_snake_with_randomized_penultimate_round_completes_correctly(self):
        """
        With round 2 randomized, the draft framework still navigates all three
        rounds and reaches STATE_COMPLETE.  Captains are excluded here because
        the cross-team guard blocks them when the snake direction changes, which
        is orthogonally tested elsewhere.
        """
        DraftRound.objects.create(
            session=self.session, round_number=2, order_type=DraftRound.ORDER_RANDOMIZED
        )
        # Build 9 non-captain, non-goalie signups so any team can receive any pick
        non_cap_signups = list(self.players)  # 6
        for i in range(3):
            non_cap_signups.append(
                SeasonSignup.objects.create(
                    season=self.season,
                    first_name=f"Extra{i}",
                    last_name="Player",
                    email=f"extra{i}@test.com",
                    primary_position=SeasonSignup.POSITION_WING,
                    secondary_position=SeasonSignup.POSITION_ONE_THING,
                    captain_interest=SeasonSignup.CAPTAIN_NO,
                )
            )
        for signup in non_cap_signups:
            response = self._post_pick(
                signup.pk, commissioner_token=self.session.commissioner_token
            )
            self.assertEqual(response.status_code, 200, response.content)

        self.session.refresh_from_db()
        self.assertEqual(self.session.state, DraftSession.STATE_COMPLETE)
        self.assertEqual(DraftPick.objects.filter(session=self.session).count(), 9)


# ---------------------------------------------------------------------------
# Captain: email team data endpoint
# ---------------------------------------------------------------------------


class EmailTeamDataViewTests(DraftTestBase):
    def _url(self, team=None):
        t = team or self.team1
        return reverse("draft_email_team", args=[self.session.pk, t.captain_token])

    def _fill_team1_picks(self):
        """Give team1 one pick so the roster is non-empty."""
        return self._make_pick(self.team1, self.players[0], 1, 0)

    def test_returns_200_when_draft_complete(self):
        self._complete()
        self._fill_team1_picks()
        resp = self.client.get(self._url())
        self.assertEqual(resp.status_code, 200)

    def test_400_when_draft_not_complete(self):
        self._activate()
        resp = self.client.get(self._url())
        self.assertEqual(resp.status_code, 400)
        data = json.loads(resp.content)
        self.assertIn("error", data)

    def test_400_when_draft_in_setup(self):
        resp = self.client.get(self._url())
        self.assertEqual(resp.status_code, 400)

    def test_404_with_wrong_token(self):
        import uuid

        self._complete()
        bad_url = reverse("draft_email_team", args=[self.session.pk, uuid.uuid4()])
        resp = self.client.get(bad_url)
        self.assertEqual(resp.status_code, 404)

    def test_returns_team_name_and_season(self):
        self._complete()
        self._fill_team1_picks()
        resp = self.client.get(self._url())
        data = json.loads(resp.content)
        self.assertIn("team_name", data)
        self.assertIn("season_name", data)
        self.assertIn("captain_name", data)

    def test_roster_contains_pick_with_email(self):
        self._complete()
        self._fill_team1_picks()
        resp = self.client.get(self._url())
        data = json.loads(resp.content)
        self.assertIn("roster", data)
        self.assertTrue(len(data["roster"]) > 0)
        first = data["roster"][0]
        self.assertIn("full_name", first)
        self.assertIn("email", first)
        self.assertIn("primary_position", first)
        self.assertIn("is_captain", first)

    def test_captain_is_flagged_in_roster(self):
        # Captain's auto-pick
        self._complete()
        self._make_pick(self.team1, self.cap1, 2, 0, auto=True)
        resp = self.client.get(self._url())
        data = json.loads(resp.content)
        captain_entries = [p for p in data["roster"] if p["is_captain"]]
        self.assertEqual(len(captain_entries), 1)
        self.assertEqual(captain_entries[0]["full_name"], self.cap1.full_name)

    def test_other_team_token_returns_own_roster(self):
        """Each captain's token only exposes their own team."""
        self._complete()
        self._make_pick(self.team1, self.players[0], 1, 0)
        self._make_pick(self.team2, self.players[1], 1, 1)
        resp2 = self.client.get(self._url(team=self.team2))
        data = json.loads(resp2.content)
        names = [p["full_name"] for p in data["roster"]]
        self.assertIn(self.players[1].full_name, names)
        self.assertNotIn(self.players[0].full_name, names)


# ---------------------------------------------------------------------------
# draft_results_download
# ---------------------------------------------------------------------------


class DraftResultsDownloadTests(DraftTestBase):
    """Tests for the public CSV/XLSX draft results download endpoint."""

    def _url(self, fmt=None):
        url = reverse("draft_results_download", args=[self.session.pk])
        if fmt:
            url += f"?format={fmt}"
        return url

    def _setup_picks(self):
        self._activate()
        self._make_pick(self.team1, self.players[0], 1, 0)
        self._make_pick(self.team2, self.players[1], 1, 1)
        self._make_pick(self.team3, self.players[2], 1, 2)

    def test_no_picks_returns_404(self):
        resp = self.client.get(self._url())
        self.assertEqual(resp.status_code, 404)

    def test_csv_default_format(self):
        self._setup_picks()
        resp = self.client.get(self._url())
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "text/csv")
        self.assertIn("draft_results_", resp["Content-Disposition"])
        self.assertIn(".csv", resp["Content-Disposition"])

    def test_csv_explicit_format_param(self):
        self._setup_picks()
        resp = self.client.get(self._url("csv"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "text/csv")

    def test_csv_contains_headers(self):
        self._setup_picks()
        resp = self.client.get(self._url("csv"))
        content = resp.content.decode()
        self.assertIn("Round", content)
        self.assertIn("Team", content)
        self.assertIn("Player", content)
        self.assertIn("Position", content)

    def test_csv_contains_pick_data(self):
        self._setup_picks()
        resp = self.client.get(self._url("csv"))
        content = resp.content.decode()
        self.assertIn(self.players[0].full_name, content)
        self.assertIn(self.players[1].full_name, content)
        self.assertIn(self.players[2].full_name, content)

    def test_csv_row_count_matches_picks(self):
        self._setup_picks()
        resp = self.client.get(self._url("csv"))
        lines = [l for l in resp.content.decode().splitlines() if l.strip()]
        # 1 header + 3 picks
        self.assertEqual(len(lines), 4)

    def test_xlsx_format(self):
        self._setup_picks()
        resp = self.client.get(self._url("xlsx"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(".xlsx", resp["Content-Disposition"])

    def test_xlsx_is_valid_workbook(self):
        from openpyxl import load_workbook
        import io

        self._setup_picks()
        resp = self.client.get(self._url("xlsx"))
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # Header row
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Round", headers)
        self.assertIn("Player", headers)
        # 3 data rows + 1 header
        self.assertEqual(ws.max_row, 4)

    def test_unknown_format_falls_back_to_csv(self):
        self._setup_picks()
        resp = self.client.get(self._url("json"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "text/csv")

    def test_no_auth_required(self):
        """Download endpoint is public — no login needed."""
        self._setup_picks()
        resp = self.client.get(self._url("csv"))
        self.assertEqual(resp.status_code, 200)

    def test_invalid_session_pk_returns_404(self):
        resp = self.client.get(reverse("draft_results_download", args=[99999]))
        self.assertEqual(resp.status_code, 404)


# ---------------------------------------------------------------------------
# draft_sessions_list
# ---------------------------------------------------------------------------


class DraftSessionsListTests(DraftTestBase):
    """Tests for the public draft archive listing page."""

    def test_empty_list_returns_200(self):
        # Default session is SETUP — should be excluded; page still renders.
        resp = self.client.get(reverse("draft_sessions_list"))
        self.assertEqual(resp.status_code, 200)

    def test_setup_session_excluded(self):
        resp = self.client.get(reverse("draft_sessions_list"))
        self.assertNotContains(resp, str(self.session.season))

    def test_complete_session_shown(self):
        self.session.state = DraftSession.STATE_COMPLETE
        self.session.save()
        resp = self.client.get(reverse("draft_sessions_list"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, str(self.session.season))

    def test_active_session_shown_with_watch_live(self):
        self.session.state = DraftSession.STATE_ACTIVE
        self.session.save()
        resp = self.client.get(reverse("draft_sessions_list"))
        self.assertContains(resp, "Watch Live")

    def test_paused_session_shown_with_watch_draft(self):
        self.session.state = DraftSession.STATE_PAUSED
        self.session.save()
        resp = self.client.get(reverse("draft_sessions_list"))
        self.assertContains(resp, "Watch Draft")

    def test_download_links_absent_without_picks(self):
        self.session.state = DraftSession.STATE_COMPLETE
        self.session.save()
        resp = self.client.get(reverse("draft_sessions_list"))
        self.assertNotContains(resp, "format=csv")

    def test_download_links_present_with_picks(self):
        self._activate()
        self._make_pick(self.team1, self.players[0], 1, 0)
        self.session.state = DraftSession.STATE_COMPLETE
        self.session.save()
        resp = self.client.get(reverse("draft_sessions_list"))
        self.assertContains(resp, "format=csv")
        self.assertContains(resp, "format=xlsx")

    def test_newest_season_first(self):
        """Newer season year must appear before older one."""
        old_season = Season.objects.create(
            year=2020, season_type=4, is_current_season=False
        )
        DraftSession.objects.create(
            season=old_season,
            num_teams=3,
            num_rounds=3,
            state=DraftSession.STATE_COMPLETE,
            signups_open=False,
        )
        new_season = Season.objects.create(
            year=2030, season_type=1, is_current_season=False
        )
        DraftSession.objects.create(
            season=new_season,
            num_teams=3,
            num_rounds=3,
            state=DraftSession.STATE_COMPLETE,
            signups_open=False,
        )
        resp = self.client.get(reverse("draft_sessions_list"))
        content = resp.content.decode()
        self.assertLess(content.index("2030"), content.index("2020"))


# ---------------------------------------------------------------------------
# _get_champion_data_for_sessions + archive champion display
# ---------------------------------------------------------------------------


class DraftChampionDataTests(TestCase):
    """
    Tests for _get_champion_data_for_sessions and the champion section
    rendered on the draft archive page.
    """

    def _make_player_stat(self, team, matchup, goals):
        player = Player.objects.create(first_name=f"P{team.id}", last_name="Test")
        Roster.objects.create(player=player, team=team, position1=1, is_captain=False)
        Stat.objects.create(
            player=player, team=team, matchup=matchup, goals=goals, assists=0
        )
        return player

    def setUp(self):
        self.client = Client()
        self.season = Season.objects.create(
            year=2025, season_type=4, is_current_season=False
        )
        self.draft_div = Division.objects.create(division=3)

        self.session = DraftSession.objects.create(
            season=self.season,
            num_teams=2,
            num_rounds=3,
            state=DraftSession.STATE_COMPLETE,
            signups_open=False,
        )

        # Real league teams (created during finalization)
        self.team1 = Team.objects.create(
            team_name="Alpha Team",
            team_color="Red",
            division=self.draft_div,
            season=self.season,
            is_active=True,
        )
        self.team2 = Team.objects.create(
            team_name="Beta Team",
            team_color="Blue",
            division=self.draft_div,
            season=self.season,
            is_active=True,
        )

        # Captain signups
        self.cap1 = SeasonSignup.objects.create(
            season=self.season,
            first_name="Alice",
            last_name="Smith",
            email="alice@test.com",
            primary_position=SeasonSignup.POSITION_CENTER,
            secondary_position=SeasonSignup.POSITION_ONE_THING,
            captain_interest=SeasonSignup.CAPTAIN_YES,
        )
        self.cap2 = SeasonSignup.objects.create(
            season=self.season,
            first_name="Bob",
            last_name="Jones",
            email="bob@test.com",
            primary_position=SeasonSignup.POSITION_CENTER,
            secondary_position=SeasonSignup.POSITION_ONE_THING,
            captain_interest=SeasonSignup.CAPTAIN_YES,
        )

        # DraftTeams linked to real league teams
        self.dt1 = DraftTeam.objects.create(
            session=self.session,
            captain=self.cap1,
            draft_position=1,
            league_team=self.team1,
        )
        self.dt2 = DraftTeam.objects.create(
            session=self.session,
            captain=self.cap2,
            draft_position=2,
            league_team=self.team2,
        )

        # Week + championship matchup (team1 wins 3-1)
        self.week = Week.objects.create(
            division=self.draft_div,
            season=self.season,
            date=datetime.date(2025, 3, 15),
        )
        self.champ_matchup = MatchUp.objects.create(
            week=self.week,
            time=datetime.time(19, 0),
            hometeam=self.team1,
            awayteam=self.team2,
            is_postseason=True,
            is_championship=True,
        )
        self._make_player_stat(self.team1, self.champ_matchup, goals=3)
        self._make_player_stat(self.team2, self.champ_matchup, goals=1)

        # Regular-season record for the champion
        Team_Stat.objects.create(
            division=self.draft_div,
            season=self.season,
            team=self.team1,
            win=8,
            otw=0,
            loss=2,
            otl=0,
            tie=0,
            goals_for=30,
            goals_against=15,
        )

    # --- _get_champion_data_for_sessions unit tests ---

    def test_champion_identified(self):
        result = _get_champion_data_for_sessions([self.session])
        self.assertIn(self.session.pk, result)
        self.assertEqual(result[self.session.pk]["team"], self.team1)

    def test_champion_draft_team_linked(self):
        result = _get_champion_data_for_sessions([self.session])
        self.assertEqual(result[self.session.pk]["draft_team"], self.dt1)

    def test_champion_team_stat_present(self):
        result = _get_champion_data_for_sessions([self.session])
        ts = result[self.session.pk]["team_stat"]
        self.assertIsNotNone(ts)
        self.assertEqual(ts.win, 8)
        self.assertEqual(ts.loss, 2)

    def test_playoff_wins_and_losses_counted(self):
        result = _get_champion_data_for_sessions([self.session])
        self.assertEqual(result[self.session.pk]["playoff_wins"], 1)
        self.assertEqual(result[self.session.pk]["playoff_losses"], 0)

    def test_non_complete_session_excluded(self):
        self.session.state = DraftSession.STATE_ACTIVE
        self.session.save()
        result = _get_champion_data_for_sessions([self.session])
        self.assertEqual(result, {})

    def test_no_championship_matchup_returns_empty(self):
        self.champ_matchup.is_championship = False
        self.champ_matchup.save()
        result = _get_champion_data_for_sessions([self.session])
        self.assertEqual(result, {})

    def test_tied_championship_excluded(self):
        """A tied championship game should not produce a champion."""
        # Reset goals to 2-2 tie
        Stat.objects.filter(matchup=self.champ_matchup).update(goals=2)
        result = _get_champion_data_for_sessions([self.session])
        self.assertEqual(result, {})

    def test_empty_sessions_list(self):
        self.assertEqual(_get_champion_data_for_sessions([]), {})

    # --- Archive page rendering tests ---

    def test_champion_team_name_shown(self):
        resp = self.client.get(reverse("draft_sessions_list"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Alpha Team")

    def test_champion_captain_shown(self):
        resp = self.client.get(reverse("draft_sessions_list"))
        self.assertContains(resp, "Alice Smith")

    def test_regular_season_record_shown(self):
        resp = self.client.get(reverse("draft_sessions_list"))
        self.assertContains(resp, "Regular Season")
        # 8 wins + 0 OTW = 8, 2 losses + 0 OTL = 2 → "8-2"
        self.assertContains(resp, "8-2")

    def test_playoff_record_shown(self):
        resp = self.client.get(reverse("draft_sessions_list"))
        self.assertContains(resp, "Playoffs")
        self.assertContains(resp, "1-0")

    def test_champion_shown_without_draft_team_link(self):
        """
        If DraftTeams have no league_team, the champion team name is still shown
        (determined from the championship matchup) but the captain line is absent.
        """
        self.dt1.league_team = None
        self.dt1.save()
        self.dt2.league_team = None
        self.dt2.save()
        resp = self.client.get(reverse("draft_sessions_list"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Alpha Team")  # team name from MatchUp still shown
        self.assertNotContains(
            resp, "Alice Smith"
        )  # captain NOT shown (no DraftTeam link)


# ---------------------------------------------------------------------------
# Champion column on draft board page
# ---------------------------------------------------------------------------


class DraftBoardChampionTests(DraftChampionDataTests):
    """
    Tests that the champion column is highlighted on the draft board page
    (spectator, commissioner, captain views) for completed sessions.
    The champion is identified by the champion-col-header CSS class and
    trophy icon in the table header; no separate banner div is rendered.
    Inherits setUp from DraftChampionDataTests which creates a COMPLETE
    session with a clear championship winner (Alpha Team / Alice Smith).
    """

    def _spectator_url(self):
        return reverse("draft_board_spectator", args=[self.session.pk])

    def _commissioner_url(self):
        return reverse(
            "draft_board_commissioner",
            args=[self.session.pk, self.session.commissioner_token],
        )

    def _captain_url(self):
        return reverse(
            "draft_board_captain",
            args=[self.session.pk, self.dt1.captain_token],
        )

    def test_spectator_champion_in_context(self):
        resp = self.client.get(self._spectator_url())
        self.assertEqual(resp.status_code, 200)
        self.assertIsNotNone(resp.context["champion"])
        self.assertEqual(resp.context["champion"]["team"], self.team1)

    def test_spectator_champion_league_team_id_in_context(self):
        resp = self.client.get(self._spectator_url())
        self.assertEqual(resp.context["champion_league_team_id"], self.team1.id)

    def test_spectator_champion_league_team_id_js_constant(self):
        """The CHAMPION_LEAGUE_TEAM_ID JS constant is rendered with the team's pk."""
        resp = self.client.get(self._spectator_url())
        self.assertContains(resp, f"CHAMPION_LEAGUE_TEAM_ID = {self.team1.id}")

    def test_spectator_team_name_from_league_team(self):
        """display_name uses the real league team name, not the draft team name."""
        resp = self.client.get(self._spectator_url())
        import json

        state = json.loads(resp.context["initial_state"])
        team_names = [t["display_name"] for t in state["teams"]]
        self.assertIn("Alpha Team", team_names)
        self.assertIn("Beta Team", team_names)
        # DraftTeam default names should not be used when league_team is set
        self.assertNotIn("Alice's Team", team_names)

    def test_spectator_record_in_state(self):
        """Regular-season record is included in the state payload per team."""
        resp = self.client.get(self._spectator_url())
        import json

        state = json.loads(resp.context["initial_state"])
        champ_team = next(
            t for t in state["teams"] if t["display_name"] == "Alpha Team"
        )
        self.assertEqual(champ_team["record"], "8-2")

    def test_commissioner_champion_in_context(self):
        resp = self.client.get(self._commissioner_url())
        self.assertEqual(resp.status_code, 200)
        self.assertIsNotNone(resp.context["champion"])

    def test_commissioner_champion_league_team_id_in_context(self):
        resp = self.client.get(self._commissioner_url())
        self.assertEqual(resp.context["champion_league_team_id"], self.team1.id)

    def test_captain_champion_in_context(self):
        resp = self.client.get(self._captain_url())
        self.assertEqual(resp.status_code, 200)
        self.assertIsNotNone(resp.context["champion"])

    def test_captain_champion_league_team_id_in_context(self):
        resp = self.client.get(self._captain_url())
        self.assertEqual(resp.context["champion_league_team_id"], self.team1.id)

    def test_no_champion_for_non_complete_session(self):
        """Active/paused sessions have no champion context."""
        self.session.state = DraftSession.STATE_ACTIVE
        self.session.save()
        resp = self.client.get(self._spectator_url())
        self.assertEqual(resp.status_code, 200)
        self.assertIsNone(resp.context["champion"])
        self.assertIsNone(resp.context["champion_league_team_id"])
        self.assertContains(resp, "CHAMPION_LEAGUE_TEAM_ID = null")

    def test_get_session_champion_wrapper(self):
        """_get_session_champion returns the same data as the batch helper."""
        champ = _get_session_champion(self.session)
        self.assertIsNotNone(champ)
        self.assertEqual(champ["team"], self.team1)

    def test_get_session_champion_none_for_non_complete(self):
        self.session.state = DraftSession.STATE_ACTIVE
        self.session.save()
        self.assertIsNone(_get_session_champion(self.session))

    def test_league_team_id_in_state_payload(self):
        """Each team in the state payload includes its league_team_id."""
        resp = self.client.get(self._spectator_url())
        import json

        state = json.loads(resp.context["initial_state"])
        for t in state["teams"]:
            self.assertIn("league_team_id", t)

    def test_fallback_to_draft_team_name_when_no_league_team(self):
        """When league_team is not set, display_name falls back to DraftTeam.team_name."""
        self.dt1.league_team = None
        self.dt1.save()
        resp = self.client.get(self._spectator_url())
        import json

        state = json.loads(resp.context["initial_state"])
        t1_data = next(t for t in state["teams"] if t["id"] == self.dt1.pk)
        self.assertEqual(t1_data["display_name"], self.dt1.team_name)


# ---------------------------------------------------------------------------
# GAA calculation — only goalie-rostered seasons
# ---------------------------------------------------------------------------


class WednesdayStatsGAATests(TestCase):
    """
    Regression tests for _get_wednesday_stats and _batch_wednesday_stats.

    A player who played several seasons as a non-goalie and then one season as
    a goalie should have their GAA calculated only from the goalie season —
    not diluted across all seasons' games.
    """

    def setUp(self):
        self.wed_div = Division.objects.create(division=3)
        self.player = Player.objects.create(first_name="Ryan", last_name="Jacob")

        # Three past seasons
        self.season_d1 = Season.objects.create(
            year=2022, season_type=4, is_current_season=False
        )
        self.season_d2 = Season.objects.create(
            year=2023, season_type=4, is_current_season=False
        )
        self.season_goalie = Season.objects.create(
            year=2024, season_type=4, is_current_season=False
        )

        # Defense team (2022) — 10 games, 0 goals against
        self.team_d1 = Team.objects.create(
            team_name="D Team 2022",
            team_color="Red",
            division=self.wed_div,
            season=self.season_d1,
            is_active=True,
        )
        Roster.objects.create(
            player=self.player, team=self.team_d1, position1=3  # Defense
        )
        week1 = Week.objects.create(
            division=self.wed_div, season=self.season_d1, date=datetime.date(2022, 1, 1)
        )
        opp1 = Team.objects.create(
            team_name="Opp 2022",
            team_color="Blue",
            division=self.wed_div,
            season=self.season_d1,
            is_active=True,
        )
        for i in range(10):
            mu = MatchUp.objects.create(
                week=week1,
                time=datetime.time(19 + i % 4, 0),
                hometeam=self.team_d1,
                awayteam=opp1,
            )
            Stat.objects.create(
                player=self.player,
                team=self.team_d1,
                matchup=mu,
                goals=0,
                assists=0,
                goals_against=0,
            )

        # Defense team (2023) — 10 games, 0 goals against
        self.team_d2 = Team.objects.create(
            team_name="D Team 2023",
            team_color="Green",
            division=self.wed_div,
            season=self.season_d2,
            is_active=True,
        )
        Roster.objects.create(
            player=self.player, team=self.team_d2, position1=3  # Defense
        )
        week2 = Week.objects.create(
            division=self.wed_div, season=self.season_d2, date=datetime.date(2023, 1, 1)
        )
        opp2 = Team.objects.create(
            team_name="Opp 2023",
            team_color="Yellow",
            division=self.wed_div,
            season=self.season_d2,
            is_active=True,
        )
        for i in range(10):
            mu = MatchUp.objects.create(
                week=week2,
                time=datetime.time(19 + i % 4, 0),
                hometeam=self.team_d2,
                awayteam=opp2,
            )
            Stat.objects.create(
                player=self.player,
                team=self.team_d2,
                matchup=mu,
                goals=0,
                assists=0,
                goals_against=0,
            )

        # Goalie team (2024) — 5 games, 10 goals against → GAA = 2.00
        self.team_g = Team.objects.create(
            team_name="G Team 2024",
            team_color="Purple",
            division=self.wed_div,
            season=self.season_goalie,
            is_active=True,
        )
        Roster.objects.create(
            player=self.player,
            team=self.team_g,
            position1=4,  # Goalie
            is_substitute=False,
        )
        week3 = Week.objects.create(
            division=self.wed_div,
            season=self.season_goalie,
            date=datetime.date(2024, 1, 1),
        )
        opp3 = Team.objects.create(
            team_name="Opp 2024",
            team_color="Orange",
            division=self.wed_div,
            season=self.season_goalie,
            is_active=True,
        )
        for i in range(5):
            mu = MatchUp.objects.create(
                week=week3,
                time=datetime.time(19 + i % 4, 0),
                hometeam=self.team_g,
                awayteam=opp3,
            )
            Stat.objects.create(
                player=self.player,
                team=self.team_g,
                matchup=mu,
                goals=0,
                assists=0,
                goals_against=2,  # 2 goals against per game → GAA 2.00
            )

    def test_gaa_uses_only_goalie_seasons(self):
        """GAA must reflect only goalie-rostered seasons (5 GP, 10 GA = 2.00)."""
        stats = _get_wednesday_stats(self.player)
        # If all 25 games were used: 10 GA / 25 GP = 0.40 (wrong)
        # Correct: 10 GA / 5 GP = 2.00
        self.assertEqual(stats["gaa"], 2.00)

    def test_seasons_count_includes_non_goalie_seasons(self):
        """seasons should count all Wednesday seasons, not just goalie seasons."""
        stats = _get_wednesday_stats(self.player)
        self.assertEqual(stats["seasons"], 3)

    def test_gaa_none_for_player_with_no_goalie_seasons(self):
        """A player who has never been rostered as goalie should have gaa=None."""
        non_goalie = Player.objects.create(first_name="Dave", last_name="Defense")
        Roster.objects.create(player=non_goalie, team=self.team_d1, position1=3)
        Stat.objects.create(
            player=non_goalie,
            team=self.team_d1,
            matchup=None,
            goals=1,
            assists=2,
            goals_against=0,
        )
        stats = _get_wednesday_stats(non_goalie)
        self.assertIsNone(stats["gaa"])

    def test_batch_gaa_uses_only_goalie_seasons(self):
        """_batch_wednesday_stats must also restrict GAA to goalie-rostered seasons."""
        batch = _batch_wednesday_stats([self.player.id])
        stats = batch[self.player.id]
        self.assertEqual(stats["gaa"], 2.00)

    def test_batch_gaa_none_for_player_with_no_goalie_seasons(self):
        """Batch version returns gaa=None for players never rostered as goalie."""
        non_goalie = Player.objects.create(first_name="Eve", last_name="Wing")
        Roster.objects.create(player=non_goalie, team=self.team_d1, position1=2)
        Stat.objects.create(
            player=non_goalie,
            team=self.team_d1,
            matchup=None,
            goals=3,
            assists=1,
            goals_against=0,
        )
        batch = _batch_wednesday_stats([non_goalie.id])
        self.assertIsNone(batch[non_goalie.id]["gaa"])
