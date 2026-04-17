"""
Views for the Wednesday Draft League signup and real-time draft board.
"""

import csv
import io
import json
import random

from django.db.models import Case, Count, Q, Sum, Value, When
from django.http import HttpResponse, JsonResponse, Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models import (
    DraftPick,
    DraftRound,
    DraftSession,
    DraftTeam,
    Division,
    Player,
    Roster,
    Season,
    SeasonSignup,
    Stat,
    Team,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _get_wednesday_stats(player):
    """
    Return career Wednesday Draft League stats for a linked Player.
    Non-goalies:  per-season averages for goals, assists, points.
    Goalies:  GAA (goals against per game played).
    New players (no recorded stats): is_new=True, all averages zero.
    Also returns adp: average round drafted across all past sessions (None if no history).
    """
    if player is None:
        return None

    WED_DIVISION = 3  # "Wednesday Draft League" in Division.DIVISION_TYPE

    qs = Stat.objects.filter(
        player=player,
        team__division__division=WED_DIVISION,
    )

    seasons_played = qs.values("team__season").distinct().count()

    # Average draft round across all past sessions this player appeared in
    past_picks = list(
        DraftPick.objects.filter(signup__linked_player=player).values_list(
            "round_number", flat=True
        )
    )
    adp_count = len(past_picks)
    adp = round(sum(past_picks) / adp_count, 1) if adp_count else None

    if seasons_played == 0:
        return {
            "is_new": True,
            "seasons": 0,
            "goals_per_season": 0,
            "assists_per_season": 0,
            "points_per_season": 0,
            "gaa": None,
            "adp": adp,
            "adp_count": adp_count,
        }

    agg = qs.aggregate(
        goals=Sum("goals"),
        assists=Sum("assists"),
        goals_against=Sum("goals_against"),
        matchup_games=Count("matchup", distinct=True),
    )

    goals = agg["goals"] or 0
    assists = agg["assists"] or 0
    goals_against = agg["goals_against"] or 0
    # Production data: each Stat row has a real MatchUp FK → count distinct matchups.
    # Seed / legacy data: matchup=None (season-total rows) → fall back to row count so
    # GAA is still calculable (each row represents one game in that case).
    games = agg["matchup_games"] or qs.count()
    points = goals + assists

    def per_season(n):
        return round(n / seasons_played, 1)

    # games > 0 is the correct guard; a shutout goalie has valid GAA of 0.00
    gaa = round(goals_against / games, 2) if games > 0 else None

    return {
        "is_new": False,
        "seasons": seasons_played,
        "goals_per_season": per_season(goals),
        "assists_per_season": per_season(assists),
        "points_per_season": per_season(points),
        "gaa": gaa,
        "adp": adp,
        "adp_count": adp_count,
    }


def _batch_wednesday_stats(player_ids):
    """
    Batch version of _get_wednesday_stats.
    Returns {player_id: stats_dict} for all given player IDs in 2 queries,
    regardless of how many players are passed.
    """
    player_ids = [p for p in player_ids if p is not None]
    if not player_ids:
        return {}

    WED_DIVISION = 3

    # Query 1: per-player season counts + aggregated stats + row count fallback
    agg_rows = {
        row["player_id"]: row
        for row in Stat.objects.filter(
            player_id__in=player_ids,
            team__division__division=WED_DIVISION,
        )
        .values("player_id")
        .annotate(
            seasons=Count("team__season", distinct=True),
            goals=Sum("goals"),
            assists=Sum("assists"),
            goals_against=Sum("goals_against"),
            matchup_games=Count("matchup", distinct=True),
            stat_rows=Count("id"),
        )
    }

    # Query 2: ADP — average round drafted across all past sessions
    adp_raw: dict = {}
    for pid, rnd in DraftPick.objects.filter(
        signup__linked_player_id__in=player_ids
    ).values_list("signup__linked_player_id", "round_number"):
        adp_raw.setdefault(pid, []).append(rnd)

    result = {}
    for pid in player_ids:
        past_picks = adp_raw.get(pid, [])
        adp_count = len(past_picks)
        adp = round(sum(past_picks) / adp_count, 1) if adp_count else None

        agg = agg_rows.get(pid)
        if not agg or not agg["seasons"]:
            result[pid] = {
                "is_new": True,
                "seasons": 0,
                "goals_per_season": 0,
                "assists_per_season": 0,
                "points_per_season": 0,
                "gaa": None,
                "adp": adp,
                "adp_count": adp_count,
            }
            continue

        seasons = agg["seasons"]
        goals = agg["goals"] or 0
        assists = agg["assists"] or 0
        goals_against = agg["goals_against"] or 0
        # Prefer distinct matchup count; fall back to row count for legacy seed data
        games = agg["matchup_games"] or agg["stat_rows"] or 0
        points = goals + assists
        gaa = round(goals_against / games, 2) if games > 0 else None

        def per_season(n, s=seasons):
            return round(n / s, 1)

        result[pid] = {
            "is_new": False,
            "seasons": seasons,
            "goals_per_season": per_season(goals),
            "assists_per_season": per_season(assists),
            "points_per_season": per_season(points),
            "gaa": gaa,
            "adp": adp,
            "adp_count": adp_count,
        }
    return result


def _signup_payload(signup, stats_cache=None):
    """
    Serialise a SeasonSignup to a dict safe for JSON / template context.
    Includes historical stats if a linked_player exists.
    Pass stats_cache (from _batch_wednesday_stats) to avoid per-player queries.
    """
    if stats_cache is not None:
        stats = (
            stats_cache.get(signup.linked_player_id)
            if signup.linked_player_id
            else None
        )
    else:
        stats = _get_wednesday_stats(signup.linked_player)
    return {
        "id": signup.pk,
        "full_name": signup.full_name,
        "first_name": signup.first_name,
        "last_name": signup.last_name,
        "is_goalie": signup.is_goalie,
        "is_returning": signup.is_returning,
        "primary_position": signup.get_primary_position_display(),
        "secondary_position": signup.get_secondary_position_display(),
        "captain_interest": signup.get_captain_interest_display(),
        "stats": stats,
    }


def _session_state_payload(session, stats_cache=None):
    """
    Full snapshot of the current draft state for broadcasting / page load.

    Pass a pre-built stats_cache (from _batch_wednesday_stats) to avoid
    recomputing it when the caller already has one (e.g. the commissioner view).
    """
    teams = list(
        session.teams.select_related("captain")
        .prefetch_related("draft_picks__signup")
        .order_by("draft_position", "pk")
    )

    drafted_signup_ids = set(
        DraftPick.objects.filter(session=session).values_list("signup_id", flat=True)
    )

    captain_signup_ids = set(session.teams.values_list("captain_id", flat=True))

    all_signups = list(
        session.season.signups.select_related("linked_player").order_by("last_name")
    )

    if stats_cache is None:
        player_ids = [s.linked_player_id for s in all_signups if s.linked_player_id]
        stats_cache = _batch_wednesday_stats(player_ids)

    available = [
        _signup_payload(s, stats_cache)
        for s in all_signups
        if s.pk not in drafted_signup_ids and s.pk not in captain_signup_ids
    ]

    teams_data = []
    for team in teams:
        picks_by_round = {}
        # A goalie-captain counts even before their auto-pick round fires.
        has_goalie = team.captain.is_goalie
        for pick in team.draft_picks.all():
            payload = _signup_payload(pick.signup, stats_cache)
            payload["is_captain_pick"] = pick.is_auto_captain
            payload["pick_id"] = pick.pk
            picks_by_round[pick.round_number] = payload
            if payload["is_goalie"]:
                has_goalie = True
        teams_data.append(
            {
                "id": team.pk,
                "team_name": team.team_name,
                "captain_name": team.captain.full_name,
                "captain_signup_id": team.captain_id,
                "draft_position": team.draft_position,
                "captain_draft_round": team.captain_draft_round,
                "picks": picks_by_round,
                "has_goalie": has_goalie,
            }
        )

    current = session.current_pick
    if current:
        cur_round, cur_pick_idx = current
        order = session.pick_order_for_round(cur_round)
        active_team_pk = order[cur_pick_idx] if cur_pick_idx < len(order) else None
    else:
        cur_round, cur_pick_idx, active_team_pk = None, None, None

    return {
        "state": session.state,
        "num_teams": session.num_teams,
        "num_rounds": session.num_rounds,
        "teams": teams_data,
        "available_players": available,
        "current_round": cur_round,
        "current_pick_index": cur_pick_idx,
        "active_team_pk": active_team_pk,
        "finalized": session.finalized_at is not None,
    }


# ---------------------------------------------------------------------------
# Public signup form
# ---------------------------------------------------------------------------


def draft_signup(request, season_pk):
    """Public signup form for a Wednesday Draft League season."""
    season = get_object_or_404(Season, pk=season_pk)

    try:
        draft_session = season.draft_session
    except DraftSession.DoesNotExist:
        raise Http404("No active signup for this season.")

    if not draft_session.signups_open:
        return render(
            request,
            "leagues/draft_signup_closed.html",
            {"season": season},
        )

    error = None
    submitted = False

    VALID_PRIMARY = {c[0] for c in SeasonSignup.PRIMARY_POSITION_CHOICES}
    VALID_SECONDARY = {c[0] for c in SeasonSignup.SECONDARY_POSITION_CHOICES}
    VALID_CAPTAIN = {c[0] for c in SeasonSignup.CAPTAIN_INTEREST_CHOICES}

    if request.method == "POST":
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        email = request.POST.get("email", "").strip()
        notes = request.POST.get("notes", "").strip()

        try:
            primary_position = int(request.POST.get("primary_position", ""))
            if primary_position not in VALID_PRIMARY:
                raise ValueError
        except (TypeError, ValueError):
            primary_position = None

        try:
            secondary_position = int(request.POST.get("secondary_position", ""))
            if secondary_position not in VALID_SECONDARY:
                raise ValueError
        except (TypeError, ValueError):
            secondary_position = None

        try:
            captain_interest = int(request.POST.get("captain_interest", ""))
            if captain_interest not in VALID_CAPTAIN:
                raise ValueError
        except (TypeError, ValueError):
            captain_interest = None

        if not first_name or not last_name:
            error = "First and last name are required."
        elif not email:
            error = "Email is required."
        elif primary_position is None:
            error = "Please select a primary position."
        elif secondary_position is None:
            error = "Please select a secondary position."
        elif captain_interest is None:
            error = "Please answer the captain interest question."

        if not error:
            # Try to link to an existing Player record
            linked_player = Player.objects.filter(
                first_name__iexact=first_name,
                last_name__iexact=last_name,
            ).first()

            existing_email = SeasonSignup.objects.filter(
                season=season,
                email__iexact=email,
            ).first()

            if existing_email:
                error = "An entry with this email address is already signed up for this season."
            else:
                SeasonSignup.objects.create(
                    season=season,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    primary_position=primary_position,
                    secondary_position=secondary_position,
                    captain_interest=captain_interest,
                    notes=notes,
                    linked_player=linked_player,
                )
                submitted = True

    context = {
        "season": season,
        "submitted": submitted,
        "error": error,
        "primary_position_choices": SeasonSignup.PRIMARY_POSITION_CHOICES,
        "secondary_position_choices": SeasonSignup.SECONDARY_POSITION_CHOICES,
        "captain_interest_choices": SeasonSignup.CAPTAIN_INTEREST_CHOICES,
    }
    return render(request, "leagues/draft_signup.html", context)


# ---------------------------------------------------------------------------
# Spectator draft board (read-only, public)
# ---------------------------------------------------------------------------


def draft_board_spectator(request, session_pk):
    """
    Public read-only view of a draft session.
    Receives live updates via WebSocket — this view just renders the shell.
    """
    session = get_object_or_404(DraftSession, pk=session_pk)

    rounds = list(session.rounds.order_by("round_number"))
    initial_state = json.dumps(_session_state_payload(session))

    context = {
        "session": session,
        "rounds": rounds,
        "initial_state": initial_state,
        "is_captain": False,
        "is_commissioner": False,
    }
    return render(request, "leagues/draft_board.html", context)


# ---------------------------------------------------------------------------
# Commissioner control view
# ---------------------------------------------------------------------------


def draft_board_commissioner(request, session_pk, token):
    """
    Commissioner's control panel for the draft.
    Full control: advance state, undo last pick, re-randomize rounds.
    """
    session = get_object_or_404(DraftSession, pk=session_pk, commissioner_token=token)
    rounds = list(session.rounds.order_by("round_number"))

    # Build stats cache once; share across state payload and all_players_json
    all_signups = list(
        session.season.signups.select_related("linked_player").order_by(
            "last_name", "first_name"
        )
    )
    stats_cache = _batch_wednesday_stats(
        [s.linked_player_id for s in all_signups if s.linked_player_id]
    )

    initial_state = json.dumps(_session_state_payload(session, stats_cache=stats_cache))
    all_players_json = json.dumps(
        [_signup_payload(s, stats_cache) for s in all_signups]
    )

    context = {
        "session": session,
        "rounds": rounds,
        "initial_state": initial_state,
        "is_captain": False,
        "is_commissioner": True,
        "commissioner_token": str(token),
        "all_players_json": all_players_json,
    }
    return render(request, "leagues/draft_board.html", context)


# ---------------------------------------------------------------------------
# Captain pick view
# ---------------------------------------------------------------------------


def draft_board_captain(request, session_pk, token):
    """
    Captain's view — same board as spectator but with pick controls
    when it's this captain's turn.
    """
    team = get_object_or_404(DraftTeam, session_id=session_pk, captain_token=token)
    session = team.session

    rounds = list(session.rounds.order_by("round_number"))
    initial_state = json.dumps(_session_state_payload(session))

    context = {
        "session": session,
        "rounds": rounds,
        "initial_state": initial_state,
        "is_captain": True,
        "is_commissioner": False,
        "captain_team_pk": team.pk,
        "captain_token": str(token),
        "captain_name": team.captain.full_name,
    }
    return render(request, "leagues/draft_board.html", context)


# ---------------------------------------------------------------------------
# Captain portal — single shareable URL listing all captain links
# ---------------------------------------------------------------------------


def draft_captain_portal(request, session_pk):
    """
    Public page listing every team's captain name and their unique draft-board
    link.  The commissioner shares this one URL with all captains; each captain
    finds their own name and clicks through to their private board.
    """
    session = get_object_or_404(DraftSession, pk=session_pk)
    teams = session.teams.select_related("captain").order_by(
        "draft_position", "captain__last_name"
    )
    return render(
        request,
        "leagues/draft_captain_portal.html",
        {"session": session, "teams": teams},
    )


# ---------------------------------------------------------------------------
# Draw phase: commissioner triggers position reveal
# ---------------------------------------------------------------------------


def draw_positions(request, session_pk, token):
    """
    Assign draft positions randomly (live reveal).
    POST-only; called by the commissioner.

    Accepts setup or draw state — if called from setup it auto-advances,
    removing the need for a separate "Start Draw Phase" button click.
    Also closes signups so no new players can register after the draw.
    """
    session = get_object_or_404(DraftSession, pk=session_pk, commissioner_token=token)

    if session.state == DraftSession.STATE_SETUP:
        session.state = DraftSession.STATE_DRAW
        session.save(update_fields=["state"])
    elif session.state != DraftSession.STATE_DRAW:
        return JsonResponse(
            {"error": "Positions can only be drawn in setup or draw phase."}, status=400
        )

    teams = list(session.teams.select_related("captain").all())
    if not teams:
        return JsonResponse({"error": "No teams configured."}, status=400)

    positions = list(range(1, len(teams) + 1))
    random.shuffle(positions)

    reveal_order = []
    for team, position in zip(teams, positions):
        team.draft_position = position
        team.save(update_fields=["draft_position"])
        reveal_order.append(
            {
                "team_pk": team.pk,
                "team_name": team.team_name,
                "captain_name": team.captain.full_name,
                "position": position,
            }
        )

    # Sort so clients can animate in position order
    reveal_order.sort(key=lambda x: x["position"])

    # Close signups now that the draw has happened
    if session.signups_open:
        session.signups_open = False
        session.save(update_fields=["signups_open"])

    # Broadcast the new DRAW state so all clients update state.state before
    # the overlay animation completes — the commissioner's panel will then
    # correctly render the captain-rounds UI when the overlay is closed.
    _broadcast_state_change(session)

    # Broadcast the reveal order separately so clients can animate the draw.
    from channels.layers import get_channel_layer
    from asgiref.sync import async_to_sync

    channel_layer = get_channel_layer()
    group_name = f"draft_{session_pk}"
    async_to_sync(channel_layer.group_send)(
        group_name,
        {
            "type": "draft.positions_drawn",
            "reveal_order": reveal_order,
        },
    )

    return JsonResponse({"reveal_order": reveal_order})


# ---------------------------------------------------------------------------
# Commissioner: advance session state
# ---------------------------------------------------------------------------


@require_POST
def advance_state(request, session_pk, token):
    """Move the session to the next logical state."""
    session = get_object_or_404(DraftSession, pk=session_pk, commissioner_token=token)

    transitions = {
        DraftSession.STATE_SETUP: DraftSession.STATE_DRAW,
        DraftSession.STATE_DRAW: DraftSession.STATE_ACTIVE,
        DraftSession.STATE_ACTIVE: DraftSession.STATE_PAUSED,
        DraftSession.STATE_PAUSED: DraftSession.STATE_ACTIVE,
    }

    next_state = transitions.get(session.state)
    if not next_state:
        return JsonResponse({"error": "No transition available."}, status=400)

    update_fields = ["state"]

    # Require every team to have a captain_draft_round before going active.
    # Captains are excluded from the available player pool and can only be
    # drafted via the auto-captain mechanism, so a missing round means they
    # would never be picked.
    if (
        next_state == DraftSession.STATE_ACTIVE
        and session.state == DraftSession.STATE_DRAW
    ):
        missing = session.teams.filter(captain_draft_round__isnull=True).select_related(
            "captain"
        )
        if missing.exists():
            names = ", ".join(t.captain.full_name for t in missing)
            return JsonResponse(
                {
                    "error": f"Captain round not set for: {names}. "
                    "Set a draft round for every captain before starting."
                },
                status=400,
            )

    # Close signups when the draft first goes active from the draw phase
    if (
        next_state == DraftSession.STATE_ACTIVE
        and session.state == DraftSession.STATE_DRAW
    ):
        session.signups_open = False
        update_fields.append("signups_open")

    session.state = next_state
    session.save(update_fields=update_fields)

    # If the draft just went active (from draw or resume), process any captain
    # auto-picks that fall on the very first slot(s).
    if next_state == DraftSession.STATE_ACTIVE:
        _process_auto_captain_picks(session)

    _broadcast_state_change(session)

    return JsonResponse({"state": _session_state_payload(session)})


# ---------------------------------------------------------------------------
# Commissioner: undo last pick
# ---------------------------------------------------------------------------


@require_POST
def undo_last_pick(request, session_pk, token):
    """Remove the most recently made pick."""
    session = get_object_or_404(DraftSession, pk=session_pk, commissioner_token=token)

    last_pick = DraftPick.objects.filter(session=session).order_by("-picked_at").first()

    if not last_pick:
        return JsonResponse({"error": "No picks to undo."}, status=400)

    was_auto_captain = last_pick.is_auto_captain
    undone = {
        "team_name": last_pick.team.team_name,
        "player_name": last_pick.signup.full_name,
        "round": last_pick.round_number,
        "pick": last_pick.pick_number + 1,
        "was_auto_captain": was_auto_captain,
    }
    last_pick.delete()

    # If we just undid an auto-captain pick, keep undoing until we reach a
    # manual pick (so the commissioner doesn't have to undo one-by-one through
    # a chain of consecutive captain slots).
    if was_auto_captain:
        while True:
            prev = (
                DraftPick.objects.filter(session=session).order_by("-picked_at").first()
            )
            if prev and prev.is_auto_captain:
                prev.delete()
            else:
                break

    _broadcast_state_change(session, extra={"undone_pick": undone})

    return JsonResponse({"undone": undone, "state": _session_state_payload(session)})


# ---------------------------------------------------------------------------
# Captain: make a pick (also commissioner can make picks on behalf)
# ---------------------------------------------------------------------------


def _process_auto_captain_picks(session):
    """
    After any pick (or state transition to active), check whether the next
    draft slot belongs to a team whose captain_draft_round matches the current
    round. If so, auto-draft the captain and keep checking (handles consecutive
    captain slots, e.g. in a snake draft where two captain rounds land back-to-back).

    Returns a list of the auto-picks created (may be empty).
    """
    auto_picks = []

    while True:
        current = session.current_pick
        if not current:
            break

        cur_round, cur_pick_idx = current
        order = session.pick_order_for_round(cur_round)
        if cur_pick_idx >= len(order):
            break

        active_team_pk = order[cur_pick_idx]
        try:
            active_team = DraftTeam.objects.select_related("captain").get(
                pk=active_team_pk
            )
        except DraftTeam.DoesNotExist:
            break

        # Only auto-pick if this round is designated as this team's captain round
        if active_team.captain_draft_round != cur_round:
            break

        captain_signup = active_team.captain

        # Safety: captain already drafted (e.g. trade moved them) — skip silently
        if DraftPick.objects.filter(session=session, signup=captain_signup).exists():
            break

        pick = DraftPick.objects.create(
            session=session,
            team=active_team,
            signup=captain_signup,
            round_number=cur_round,
            pick_number=cur_pick_idx,
            is_auto_captain=True,
        )
        auto_picks.append(pick)

    # If any auto-picks were made, check for draft completion
    if auto_picks:
        total_picks = session.picks.count()
        expected = session.num_teams * session.num_rounds
        if total_picks >= expected:
            session.state = DraftSession.STATE_COMPLETE
            session.save(update_fields=["state"])

    return auto_picks


@require_POST
def make_pick(request, session_pk):
    """
    Submit a draft pick.
    Either captain_token or commissioner_token must be in POST data.
    """
    session = get_object_or_404(DraftSession, pk=session_pk)

    if session.state != DraftSession.STATE_ACTIVE:
        return JsonResponse({"error": "Draft is not active."}, status=400)

    captain_token = request.POST.get("captain_token")
    commissioner_token = request.POST.get("commissioner_token")
    signup_pk = request.POST.get("signup_pk")

    # Determine who is picking
    is_commissioner = (
        commissioner_token and str(session.commissioner_token) == commissioner_token
    )
    picking_team = None

    if captain_token:
        picking_team = (
            DraftTeam.objects.select_related("captain")
            .filter(session=session, captain_token=captain_token)
            .first()
        )
        if not picking_team:
            return JsonResponse({"error": "Invalid captain token."}, status=403)
    elif not is_commissioner:
        return JsonResponse({"error": "Authentication required."}, status=403)

    # Determine whose turn it is
    current = session.current_pick
    if not current:
        return JsonResponse({"error": "Draft is complete."}, status=400)

    cur_round, cur_pick_idx = current
    order = session.pick_order_for_round(cur_round)
    active_team_pk = order[cur_pick_idx]

    if picking_team and picking_team.pk != active_team_pk:
        return JsonResponse({"error": "It is not your turn."}, status=400)

    if not picking_team:
        picking_team = DraftTeam.objects.select_related("captain").get(
            pk=active_team_pk
        )

    # If this slot is the team's designated captain auto-draft round and the captain
    # has not yet been picked, fire the auto-pick regardless of what was submitted.
    # This guards against manual picks sneaking through when _process_auto_captain_picks
    # hasn't run (e.g. after a server restart with an already-active draft).
    if (
        picking_team.captain_draft_round == cur_round
        and not DraftPick.objects.filter(
            session=session, signup=picking_team.captain
        ).exists()
    ):
        _process_auto_captain_picks(session)
        _broadcast_state_change(session, prev_round=cur_round)
        return JsonResponse({"success": True, "state": _session_state_payload(session)})

    # Validate the signup
    try:
        signup = SeasonSignup.objects.get(pk=signup_pk, season=session.season)
    except SeasonSignup.DoesNotExist:
        return JsonResponse({"error": "Player not found."}, status=400)

    if DraftPick.objects.filter(session=session, signup=signup).exists():
        return JsonResponse({"error": "Player already drafted."}, status=400)

    # Captains may only be drafted by their own team
    captain_team = session.teams.filter(captain=signup).first()
    if captain_team and captain_team.pk != picking_team.pk:
        return JsonResponse(
            {
                "error": f"{signup.full_name} is the captain of {captain_team.team_name} and can only be on their own team."
            },
            status=400,
        )

    if signup.primary_position == SeasonSignup.POSITION_GOALIE:
        # A goalie-captain counts as the team's goalie even before their
        # auto-pick round fires, so block any additional goalie pick.
        captain_is_goalie = (
            picking_team.captain.primary_position == SeasonSignup.POSITION_GOALIE
        )
        team_already_has_goalie = (
            captain_is_goalie
            or DraftPick.objects.filter(
                session=session,
                team=picking_team,
                signup__primary_position=SeasonSignup.POSITION_GOALIE,
            ).exists()
        )
        if team_already_has_goalie:
            return JsonResponse(
                {"error": f"{picking_team.team_name} already has a goalie."},
                status=400,
            )

    pick = DraftPick.objects.create(
        session=session,
        team=picking_team,
        signup=signup,
        round_number=cur_round,
        pick_number=cur_pick_idx,
    )

    # Check if draft is now complete
    total_picks = session.picks.count()
    expected = session.num_teams * session.num_rounds
    if total_picks >= expected:
        session.state = DraftSession.STATE_COMPLETE
        session.save(update_fields=["state"])
    else:
        # Advance through any consecutive captain auto-pick slots
        _process_auto_captain_picks(session)

    _broadcast_state_change(
        session,
        extra={
            "pick": {
                "team_pk": picking_team.pk,
                "team_name": picking_team.team_name,
                "player": _signup_payload(signup),
                "round": cur_round,
                "pick_number": cur_pick_idx + 1,
            }
        },
        prev_round=cur_round,
    )

    return JsonResponse({"success": True, "state": _session_state_payload(session)})


# ---------------------------------------------------------------------------
# Internal: broadcast helpers
# ---------------------------------------------------------------------------


def _broadcast_state_change(session, extra=None, prev_round=None):
    """Push a full state refresh to all connected WebSocket clients.

    prev_round: the round number that was active *before* the triggering pick
    (including any auto-captain picks processed afterwards).  When provided,
    a round-boundary crossing can be detected even if auto-captain picks
    consumed slot 0 of the new round before the broadcast ran.
    """
    from asgiref.sync import async_to_sync
    from channels.layers import get_channel_layer

    state_payload = _session_state_payload(session)
    payload = {"type": "draft.state_update", "state": state_payload}

    cur_round = state_payload["current_round"]

    # Detect whether we just crossed into a new round.  Prefer the explicit
    # prev_round signal (handles auto-captain picks at slot 0); fall back to
    # the pick-index check for callers that don't supply it (advance_state,
    # undo, etc.).
    just_entered_new_round = (
        prev_round is not None and cur_round is not None and prev_round != cur_round
    )
    at_round_start = state_payload["current_pick_index"] == 0

    if (
        state_payload["state"] == DraftSession.STATE_ACTIVE
        and cur_round is not None
        and (just_entered_new_round or at_round_start)
    ):
        cur_round = state_payload["current_round"]
        try:
            round_obj = session.rounds.get(round_number=cur_round)
            if round_obj.order_type == DraftRound.ORDER_RANDOMIZED:
                order_pks = session.pick_order_for_round(cur_round)
                teams_by_pk = {
                    t.pk: t for t in session.teams.select_related("captain").all()
                }
                payload["randomized_round_reveal"] = {
                    "round": cur_round,
                    "order": [
                        {
                            "position": i + 1,
                            "team_name": teams_by_pk[pk].team_name,
                            "captain_name": teams_by_pk[pk].captain.full_name,
                        }
                        for i, pk in enumerate(order_pks)
                        if pk in teams_by_pk
                    ],
                }
        except DraftRound.DoesNotExist:
            pass

    if extra:
        payload.update(extra)

    channel_layer = get_channel_layer()
    group_name = f"draft_{session.pk}"
    async_to_sync(channel_layer.group_send)(group_name, payload)


# ---------------------------------------------------------------------------
# Commissioner: finalize draft → create Team and Roster records
# ---------------------------------------------------------------------------


@require_POST
def finalize_draft(request, session_pk, token):
    """
    Convert draft picks into real Team and Roster records.

    First call:  creates Team objects and builds Roster from picks.
    Re-finalize: keeps existing Teams, wipes their Rosters and rebuilds from
                 current picks — so edits made via swap_pick are reflected.
    """
    from django.db import transaction

    session = get_object_or_404(DraftSession, pk=session_pk, commissioner_token=token)

    if session.state != DraftSession.STATE_COMPLETE:
        return JsonResponse(
            {"error": "Draft must be complete before finalizing."}, status=400
        )

    wed_division, _ = Division.objects.get_or_create(division=3)

    draft_teams = list(
        session.teams.select_related("captain").order_by("draft_position", "pk")
    )

    created_teams = 0
    created_rosters = 0

    with transaction.atomic():
        for draft_team in draft_teams:
            # Create the real Team once; keep it on re-finalize
            if draft_team.league_team_id is not None:
                league_team = draft_team.league_team
            else:
                league_team = Team.objects.create(
                    team_name=draft_team.team_name,
                    season=session.season,
                    division=wed_division,
                    team_color="#8b1a1a",
                    is_active=True,
                    team_photo=None,
                )
                draft_team.league_team = league_team
                draft_team.save(update_fields=["league_team"])
                created_teams += 1

            # Always wipe and rebuild rosters so edits are reflected
            Roster.objects.filter(team=league_team).delete()

            for pick in draft_team.draft_picks.select_related(
                "signup", "signup__linked_player"
            ).all():
                signup = pick.signup

                # Resolve or create the Player record
                if signup.linked_player:
                    player = signup.linked_player
                else:
                    player, _ = Player.objects.get_or_create(
                        first_name=signup.first_name,
                        last_name=signup.last_name,
                    )
                    signup.linked_player = player
                    signup.save(update_fields=["linked_player"])

                pos2 = (
                    signup.secondary_position
                    if signup.secondary_position != SeasonSignup.POSITION_ONE_THING
                    else None
                )

                Roster.objects.create(
                    player=player,
                    team=league_team,
                    position1=signup.primary_position,
                    position2=pos2,
                    is_captain=(signup.pk == draft_team.captain_id),
                    is_primary_goalie=signup.is_goalie,
                )
                created_rosters += 1

        session.finalized_at = timezone.now()
        session.save(update_fields=["finalized_at"])

    return JsonResponse(
        {
            "success": True,
            "created_teams": created_teams,
            "created_rosters": created_rosters,
            "state": _session_state_payload(session),
        }
    )


@require_POST
def swap_pick(request, session_pk, token):
    """
    Swap the player assigned to one pick slot with another player.

    If the target player is already in another pick, their slots are exchanged.
    If they are undrafted, they simply replace the current pick.
    Requires the draft to be complete.
    """
    from django.db import transaction

    session = get_object_or_404(DraftSession, pk=session_pk, commissioner_token=token)

    if session.state not in (DraftSession.STATE_COMPLETE, DraftSession.STATE_PAUSED):
        return JsonResponse(
            {"error": "Can only edit picks when the draft is paused or complete."},
            status=400,
        )

    pick_id = request.POST.get("pick_id")
    new_signup_pk = request.POST.get("new_signup_pk")

    try:
        pick_a = DraftPick.objects.get(pk=pick_id, session=session)
    except DraftPick.DoesNotExist:
        return JsonResponse({"error": "Pick not found."}, status=404)

    try:
        new_signup = SeasonSignup.objects.get(pk=new_signup_pk, season=session.season)
    except SeasonSignup.DoesNotExist:
        return JsonResponse({"error": "Player not found."}, status=404)

    if pick_a.signup_id == new_signup.pk:
        return JsonResponse({"success": True, "state": _session_state_payload(session)})

    # Captains may only be on their own team
    captain_team = session.teams.filter(captain=new_signup).first()
    if captain_team and captain_team.pk != pick_a.team_id:
        return JsonResponse(
            {
                "error": f"{new_signup.full_name} is the captain of {captain_team.team_name} and can only be on their own team."
            },
            status=400,
        )

    with transaction.atomic():
        # If the target player is already in a different pick, swap their slots.
        pick_b = (
            DraftPick.objects.filter(session=session, signup=new_signup)
            .exclude(pk=pick_a.pk)
            .first()
        )

        old_signup_id = pick_a.signup_id

        if pick_b:
            # Delete pick_b first so there is no momentary duplicate signup
            # (SQLite checks unique constraints per-row even within a single
            # UPDATE statement, so sequential saves or a CASE UPDATE both fail).
            pick_b_team_id = pick_b.team_id
            pick_b_round = pick_b.round_number
            pick_b_number = pick_b.pick_number
            pick_b.delete()

            pick_a.signup_id = new_signup.pk
            pick_a.is_auto_captain = False
            pick_a.save(update_fields=["signup_id", "is_auto_captain"])

            DraftPick.objects.create(
                session=session,
                team_id=pick_b_team_id,
                signup_id=old_signup_id,
                round_number=pick_b_round,
                pick_number=pick_b_number,
                is_auto_captain=False,
            )
        else:
            pick_a.signup = new_signup
            pick_a.is_auto_captain = False
            pick_a.save(update_fields=["signup", "is_auto_captain"])

    _broadcast_state_change(session)
    return JsonResponse({"success": True, "state": _session_state_payload(session)})


# ---------------------------------------------------------------------------
# Commissioner: set captain draft rounds (during setup or draw phase)
# ---------------------------------------------------------------------------


@require_POST
def set_captain_rounds(request, session_pk, token):
    """
    Set the captain_draft_round for each team.  Only allowed before the draft
    goes active (setup or draw state).

    POST body (JSON): {"rounds": {"<team_pk>": <round_number_or_null>, ...}}
    """
    import json as _json

    session = get_object_or_404(DraftSession, pk=session_pk, commissioner_token=token)

    if session.state not in (DraftSession.STATE_SETUP, DraftSession.STATE_DRAW):
        return JsonResponse(
            {"error": "Captain rounds can only be set before the draft starts."},
            status=400,
        )

    try:
        body = _json.loads(request.body)
        rounds_map = body.get("rounds", {})
    except (_json.JSONDecodeError, AttributeError):
        return JsonResponse({"error": "Invalid JSON body."}, status=400)

    teams = {str(t.pk): t for t in session.teams.all()}
    num_rounds = session.num_rounds

    for team_pk_str, round_val in rounds_map.items():
        team = teams.get(team_pk_str)
        if not team:
            continue
        if round_val is None:
            team.captain_draft_round = None
        else:
            try:
                r = int(round_val)
            except (TypeError, ValueError):
                return JsonResponse(
                    {"error": f"Invalid round value for team {team_pk_str}."},
                    status=400,
                )
            if r < 1 or r > num_rounds:
                return JsonResponse(
                    {"error": f"Round {r} is out of range (1–{num_rounds})."},
                    status=400,
                )
            team.captain_draft_round = r
        team.save(update_fields=["captain_draft_round"])

    _broadcast_state_change(session)
    return JsonResponse({"success": True, "state": _session_state_payload(session)})


# ---------------------------------------------------------------------------
# Captain: email-team data
# ---------------------------------------------------------------------------


def email_team_data(request, session_pk, token):
    """
    Return the captain's team roster with player emails so the client can
    build a mailto: link.  Only accessible via the captain's private token.
    Draft must be complete.
    """
    captain_team = get_object_or_404(
        DraftTeam, session_id=session_pk, captain_token=token
    )
    session = captain_team.session

    if session.state != DraftSession.STATE_COMPLETE:
        return JsonResponse(
            {"error": "Team email is only available after the draft is complete."},
            status=400,
        )

    picks = captain_team.draft_picks.select_related("signup").order_by(
        "round_number", "pick_number"
    )

    roster = []
    for pick in picks:
        s = pick.signup
        roster.append(
            {
                "full_name": s.full_name,
                "email": s.email,
                "primary_position": s.primary_position,
                "is_captain": s.pk == captain_team.captain_id,
            }
        )

    return JsonResponse(
        {
            "team_name": captain_team.team_name,
            "season_name": str(session.season),
            "captain_name": captain_team.captain.full_name,
            "roster": roster,
        }
    )


# ---------------------------------------------------------------------------
# Commissioner: reset draft back to pre-draw setup
# ---------------------------------------------------------------------------


@require_POST
def reset_draft(request, session_pk, token):
    """
    Wipe all picks and draft positions for this session, returning it to
    the SETUP state.  Picks are permanently deleted so they do not affect
    historical ADP calculations.
    """
    from django.db import transaction

    session = get_object_or_404(DraftSession, pk=session_pk, commissioner_token=token)

    if session.state == DraftSession.STATE_SETUP:
        return JsonResponse({"error": "Draft has not been started yet."}, status=400)

    with transaction.atomic():
        # Remove all picks — these must not count towards ADP
        DraftPick.objects.filter(session=session).delete()

        # Clear draft positions and any league_team links on every team slot
        session.teams.update(draft_position=None, league_team=None)

        # Return session to pre-draw state
        session.state = DraftSession.STATE_SETUP
        session.signups_open = True
        session.finalized_at = None
        session.save(update_fields=["state", "signups_open", "finalized_at"])

    _broadcast_state_change(session)
    return JsonResponse({"success": True, "state": _session_state_payload(session)})


# ---------------------------------------------------------------------------
# Draft results download (CSV / XLSX)
# ---------------------------------------------------------------------------

_DOWNLOAD_HEADERS = [
    "Round",
    "Pick #",
    "Team",
    "Captain",
    "Player",
    "Position",
    "Goals/Season",
    "Assists/Season",
    "Points/Season",
    "GAA",
    "ADP",
]


def _draft_results_rows(session):
    """
    Return list of rows for the draft results table.
    One row per pick, ordered by round then pick number.
    """
    picks = (
        DraftPick.objects.filter(session=session)
        .select_related("team__captain", "signup__linked_player")
        .order_by("round_number", "pick_number")
    )

    rows = []
    for pick in picks:
        stats = _get_wednesday_stats(pick.signup.linked_player)
        rows.append(
            [
                pick.round_number,
                pick.pick_number,
                pick.team.team_name,
                pick.team.captain.full_name,
                pick.signup.full_name,
                pick.signup.get_primary_position_display(),
                stats["goals_per_season"]
                if stats and not pick.signup.is_goalie
                else "",
                stats["assists_per_season"]
                if stats and not pick.signup.is_goalie
                else "",
                stats["points_per_season"]
                if stats and not pick.signup.is_goalie
                else "",
                stats["gaa"] if stats and pick.signup.is_goalie else "",
                stats["adp"] if stats else "",
            ]
        )
    return rows


def draft_results_download(request, session_pk):
    """
    Public download of the draft results as CSV or XLSX.
    Query param: ?format=xlsx  (default: csv)
    Only available once picks exist.
    """
    session = get_object_or_404(DraftSession, pk=session_pk)

    if not DraftPick.objects.filter(session=session).exists():
        return HttpResponse("No picks have been made yet.", status=404)

    fmt = request.GET.get("format", "csv").lower()
    rows = _draft_results_rows(session)
    season_label = str(session.season).replace(" ", "_")
    filename_base = f"draft_results_{season_label}"

    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Draft Results"

        # Header row
        ws.append(_DOWNLOAD_HEADERS)
        header_fill = PatternFill("solid", fgColor="2B6CB0")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows — alternate shading by team block
        prev_team = None
        shade = False
        shading_fills = [
            PatternFill("solid", fgColor="EBF4FF"),
            PatternFill("solid", fgColor="FFFFFF"),
        ]
        for row in rows:
            ws.append(row)
            team_name = row[2]
            if team_name != prev_team:
                shade = not shade
                prev_team = team_name
            fill = shading_fills[0] if shade else shading_fills[1]
            for cell in ws[ws.max_row]:
                cell.fill = fill

        # Column widths
        col_widths = [7, 7, 22, 20, 22, 12, 13, 15, 14, 8, 8]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
        return response

    # Default: CSV
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
    writer = csv.writer(response)
    writer.writerow(_DOWNLOAD_HEADERS)
    writer.writerows(rows)
    return response


# ---------------------------------------------------------------------------
# Public draft archive (listing of all past/current draft sessions)
# ---------------------------------------------------------------------------


def draft_sessions_list(request):
    """
    Public listing of all draft sessions, newest first.
    SETUP-state sessions are hidden — the draft becomes visible to fans
    once the commissioner advances to the draw phase.
    """
    sessions = (
        DraftSession.objects.select_related("season")
        .exclude(state=DraftSession.STATE_SETUP)
        .annotate(pick_count=Count("picks"))
        .order_by("-season__year", "-season__season_type")
    )
    return render(request, "leagues/draft_archive.html", {"sessions": sessions})
